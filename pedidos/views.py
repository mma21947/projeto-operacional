from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from .models import Pedido, ItemPedido, PedidoLog
from produtos.models import Produto
from escolas.models import Escola, Supervisor
from django.utils import timezone
from django.db import transaction
from django.contrib.auth.decorators import login_required, permission_required
from django.conf import settings
import os
from django.db.models import Q
from datetime import datetime, timedelta
from django.db.models import Count, Sum, Avg, F, Value, ExpressionWrapper, DecimalField
from django.db.models.functions import TruncMonth, TruncWeek, Extract

@login_required
@permission_required('pedidos.view_pedido')
def lista(request):
    """Exibe a lista de pedidos"""
    status_filtro = request.GET.get('status')
    
    # Base query para filtrar pelo status se especificado
    if status_filtro:
        pedidos_query = Pedido.objects.filter(status=status_filtro)
    else:
        pedidos_query = Pedido.objects.all()
    
    # Verifica se o usuário é um supervisor (Operador de Pedidos)
    if request.user.groups.filter(name='Operador de Pedidos').exists():
        try:
            # Obtém o supervisor associado ao usuário logado
            supervisor = Supervisor.objects.get(usuario=request.user)
            
            # Filtra pedidos apenas das escolas que o supervisor gerencia
            pedidos = pedidos_query.filter(escola__supervisor=supervisor)
            
            # Mensagem informativa
            messages.info(request, f'Você está visualizando apenas pedidos de contratos sob sua supervisão.')
            
        except Supervisor.DoesNotExist:
            # Supervisor não encontrado, não mostra pedidos
            pedidos = Pedido.objects.none()
            messages.warning(request, 'Sua conta de usuário não está corretamente vinculada a um supervisor. Por favor, contate o administrador.')
    else:
        # Usuário não é supervisor, mostra todos os pedidos conforme a query base
        pedidos = pedidos_query
        
    # Ordenação por data de solicitação
    pedidos = pedidos.order_by('-data_solicitacao')
    
    return render(request, 'pedidos/lista.html', {
        'pedidos': pedidos,
        'status_filtro': status_filtro
    })

@login_required
@permission_required('pedidos.add_pedido')
def novo(request):
    """Cria um novo pedido"""
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Extrair dados básicos do pedido
                escola_id = request.POST.get('escola')
                # Sempre usar a data e hora atual, independente do valor enviado pelo formulário
                data_solicitacao = timezone.now()
                observacoes = request.POST.get('observacoes', '')
                
                # Validar dados
                if not escola_id:
                    messages.error(request, "Selecione uma escola para o pedido.")
                    
                    # Filtrar escolas conforme o usuário
                    if request.user.groups.filter(name='Operador de Pedidos').exists():
                        try:
                            supervisor = Supervisor.objects.get(usuario=request.user)
                            escolas_query = Escola.objects.filter(supervisor=supervisor, ativo=True)
                            
                            # Filtrar escolas com orçamento vencido
                            data_atual = timezone.localdate()
                            escolas = [escola for escola in escolas_query if 
                                     not escola.data_validade_budget or  # Sem data de validade (válido para sempre)
                                     escola.data_validade_budget >= data_atual]  # Data de validade maior ou igual à atual
                            
                            # Mensagem informativa
                            if not escolas_query.exists():
                                messages.warning(request, 'Você não possui contratos sob sua supervisão. Entre em contato com o administrador.')
                            elif len(escolas) < escolas_query.count():
                                messages.info(request, 'Alguns contratos sob sua supervisão possuem orçamento vencido e foram ocultados da lista.')
                            else:
                                messages.info(request, f'Você está criando um pedido para os contratos sob sua supervisão.')
                                
                        except Supervisor.DoesNotExist:
                            escolas = []
                            messages.warning(request, 'Sua conta de usuário não está corretamente vinculada a um supervisor. Por favor, contate o administrador.')
                    else:
                        escolas_query = Escola.objects.filter(ativo=True)
                        
                        # Filtrar escolas com orçamento vencido
                        data_atual = timezone.localdate()
                        escolas = [escola for escola in escolas_query if 
                                 not escola.data_validade_budget or  # Sem data de validade (válido para sempre)
                                 escola.data_validade_budget >= data_atual]  # Data de validade maior ou igual à atual
                        
                        if len(escolas) < escolas_query.count():
                            messages.info(request, 'Alguns contratos com orçamento vencido foram ocultados da lista.')
                        
                        # Ordenar a lista de escolas por nome
                        escolas.sort(key=lambda escola: escola.nome)
                        
                    produtos = Produto.objects.filter(ativo=True).order_by('nome')
                    return render(request, 'pedidos/form.html', {
                        'pedido': None,
                        'escolas': escolas,
                        'produtos': produtos,
                        'status_choices': Pedido.STATUS_CHOICES
                    })
                
                # Para supervisores, verificar se a escola selecionada está sob sua supervisão
                if request.user.groups.filter(name='Operador de Pedidos').exists():
                    try:
                        supervisor = Supervisor.objects.get(usuario=request.user)
                        escola = Escola.objects.get(id=escola_id)
                        
                        if escola.supervisor != supervisor:
                            messages.error(request, "Você não tem permissão para criar pedidos para esta escola.")
                            escolas = Escola.objects.filter(supervisor=supervisor, ativo=True).order_by('nome')
                            produtos = Produto.objects.filter(ativo=True).order_by('nome')
                            return render(request, 'pedidos/form.html', {
                                'pedido': None,
                                'escolas': escolas,
                                'produtos': produtos,
                                'status_choices': Pedido.STATUS_CHOICES
                            })
                    except (Supervisor.DoesNotExist, Escola.DoesNotExist):
                        messages.error(request, "Erro ao verificar permissões.")
                        return redirect('pedidos:lista')
                
                # Criar o pedido
                pedido = Pedido(
                    escola_id=escola_id,
                    data_solicitacao=data_solicitacao,
                    observacoes=observacoes,
                    status='pendente'
                )
                pedido.save()
                
                # Registrar log de criação
                PedidoLog.objects.create(
                    pedido=pedido,
                    usuario=request.user,
                    acao='criacao',
                    status_novo='pendente',
                    descricao=f"Pedido criado por {request.user.get_full_name() or request.user.username}"
                )
                
                # Processar itens do pedido
                produtos_ids = request.POST.getlist('produto[]')
                quantidades = request.POST.getlist('quantidade[]')
                
                # Validar se há itens
                if not produtos_ids or len(produtos_ids) == 0:
                    messages.error(request, "O pedido deve ter pelo menos um item.")
                    pedido.delete()  # Remover pedido sem itens
                    
                    # Filtrar escolas conforme o usuário
                    if request.user.groups.filter(name='Operador de Pedidos').exists():
                        try:
                            supervisor = Supervisor.objects.get(usuario=request.user)
                            escolas = Escola.objects.filter(supervisor=supervisor, ativo=True).order_by('nome')
                        except Supervisor.DoesNotExist:
                            escolas = Escola.objects.none()
                    else:
                        escolas = Escola.objects.filter(ativo=True).order_by('nome')
                        
                    produtos = Produto.objects.filter(ativo=True).order_by('nome')
                    return render(request, 'pedidos/form.html', {
                        'pedido': None,
                        'escolas': escolas,
                        'produtos': produtos,
                        'status_choices': Pedido.STATUS_CHOICES
                    })
                
                # Criar os itens do pedido e calcular o valor total do pedido atual
                valor_total_atual = 0
                for i in range(len(produtos_ids)):
                    if not produtos_ids[i]:  # Pula itens sem produto selecionado
                        continue
                        
                    try:
                        quantidade = int(quantidades[i])
                        if quantidade <= 0:
                            continue
                    except (ValueError, IndexError):
                        continue
                    
                    # Buscar produto e seu valor atual
                    try:
                        produto = Produto.objects.get(id=produtos_ids[i])
                        
                        # Calcular valor do item
                        valor_item = quantidade * produto.valor_unitario
                        valor_total_atual += valor_item
                        
                        # Criar o item do pedido
                        ItemPedido.objects.create(
                            pedido=pedido,
                            produto=produto,
                            quantidade=quantidade,
                            valor_unitario=produto.valor_unitario
                        )
                    except Produto.DoesNotExist:
                        continue
                
                # Verificar se algum item foi adicionado
                itens_count = ItemPedido.objects.filter(pedido=pedido).count()
                if itens_count == 0:
                    messages.error(request, "Nenhum item válido foi adicionado ao pedido.")
                    pedido.delete()
                    
                    # Filtrar escolas conforme o usuário
                    if request.user.groups.filter(name='Operador de Pedidos').exists():
                        try:
                            supervisor = Supervisor.objects.get(usuario=request.user)
                            escolas = Escola.objects.filter(supervisor=supervisor, ativo=True).order_by('nome')
                        except Supervisor.DoesNotExist:
                            escolas = Escola.objects.none()
                    else:
                        escolas = Escola.objects.filter(ativo=True).order_by('nome')
                        
                    produtos = Produto.objects.filter(ativo=True).order_by('nome')
                    return render(request, 'pedidos/form.html', {
                        'pedido': None,
                        'escolas': escolas,
                        'produtos': produtos,
                        'status_choices': Pedido.STATUS_CHOICES
                    })
                
                # Verificar se o valor total dos pedidos da escola não excede o orçamento
                escola = Escola.objects.get(id=escola_id)
                
                # Verificar se o orçamento está dentro da validade
                if escola.budget > 0 and hasattr(escola, 'data_validade_budget') and escola.data_validade_budget:
                    data_atual = timezone.localdate()
                    if escola.data_validade_budget < data_atual:
                        # Remover o pedido criado pois o orçamento expirou
                        pedido.delete()
                        
                        messages.error(
                            request, 
                            "⚠️ ORÇAMENTO EXPIRADO! Não é possível cadastrar pedidos neste contrato pois a data de validade do orçamento já passou."
                        )
                        
                        # Filtrar escolas conforme o usuário
                        if request.user.groups.filter(name='Operador de Pedidos').exists():
                            try:
                                supervisor = Supervisor.objects.get(usuario=request.user)
                                escolas = Escola.objects.filter(supervisor=supervisor, ativo=True).order_by('nome')
                            except Supervisor.DoesNotExist:
                                escolas = Escola.objects.none()
                        else:
                            escolas = Escola.objects.filter(ativo=True).order_by('nome')
                            
                        produtos = Produto.objects.filter(ativo=True).order_by('nome')
                        return render(request, 'pedidos/form.html', {
                            'pedido': None,
                            'escolas': escolas,
                            'produtos': produtos,
                            'status_choices': Pedido.STATUS_CHOICES
                        })
                
                # Verificamos apenas se o budget da escola está definido e válido
                if escola.budget > 0:
                    # Calcular o valor total de todos os pedidos ativos da escola (excluindo cancelados)
                    pedidos_anteriores = Pedido.objects.filter(
                        escola=escola, 
                        status__in=['pendente', 'aprovado', 'pedido_enviado', 'entregue']
                    ).exclude(id=pedido.id)  # Excluir o pedido atual que já foi criado
                    
                    # Calcular o valor total dos pedidos anteriores manualmente
                    valor_total_anterior = 0
                    for p in pedidos_anteriores:
                        itens_pedido = ItemPedido.objects.filter(pedido=p)
                        for item in itens_pedido:
                            valor_total_anterior += item.quantidade * item.valor_unitario
                    
                    valor_total_com_atual = valor_total_anterior + valor_total_atual
                    
                    if valor_total_com_atual > escola.budget:
                        # Remover o pedido criado pois excede o orçamento
                        pedido.delete()
                        
                        messages.error(
                            request, 
                            "⚠️ LIMITE DE ORÇAMENTO EXCEDIDO! Não é possível cadastrar este pedido pois ultrapassa o valor disponível para este contrato."
                        )
                        
                        # Filtrar escolas conforme o usuário
                        if request.user.groups.filter(name='Operador de Pedidos').exists():
                            try:
                                supervisor = Supervisor.objects.get(usuario=request.user)
                                escolas = Escola.objects.filter(supervisor=supervisor, ativo=True).order_by('nome')
                            except Supervisor.DoesNotExist:
                                escolas = Escola.objects.none()
                        else:
                            escolas = Escola.objects.filter(ativo=True).order_by('nome')
                            
                        produtos = Produto.objects.filter(ativo=True).order_by('nome')
                        return render(request, 'pedidos/form.html', {
                            'pedido': None,
                            'escolas': escolas,
                            'produtos': produtos,
                            'status_choices': Pedido.STATUS_CHOICES
                        })
                
                messages.success(request, "Pedido criado com sucesso!")
                return redirect('pedidos:detalhes', pk=pedido.pk)
                
        except Exception as e:
            messages.error(request, f"Erro ao salvar o pedido: {str(e)}")
    
    # Filtrar escolas conforme o usuário
    if request.user.groups.filter(name='Operador de Pedidos').exists():
        try:
            supervisor = Supervisor.objects.get(usuario=request.user)
            escolas_query = Escola.objects.filter(supervisor=supervisor, ativo=True)
            
            # Filtrar escolas com orçamento vencido
            data_atual = timezone.localdate()
            
            # Verificar se existem escolas com orçamento vencido
            escolas_vencidas = [escola for escola in escolas_query if 
                              escola.data_validade_budget and  # Tem data de validade
                              escola.data_validade_budget < data_atual]  # Data de validade menor que atual
            
            # Filtrar escolas válidas
            escolas = [escola for escola in escolas_query if 
                     not escola.data_validade_budget or  # Sem data de validade (válido para sempre)
                     escola.data_validade_budget >= data_atual]  # Data de validade maior ou igual à atual
            
            # Mensagem informativa
            if not escolas_query.exists():
                messages.warning(request, 'Você não possui contratos sob sua supervisão. Entre em contato com o administrador.')
            elif escolas_vencidas:
                messages.info(request, 'Alguns contratos sob sua supervisão possuem orçamento vencido e foram ocultados da lista.')
            else:
                messages.info(request, f'Você está criando um pedido para os contratos sob sua supervisão.')
                
        except Supervisor.DoesNotExist:
            escolas = []
            messages.warning(request, 'Sua conta de usuário não está corretamente vinculada a um supervisor. Por favor, contate o administrador.')
    else:
        # Usuários administradores ou outros veem todas as escolas
        escolas_query = Escola.objects.filter(ativo=True)
        
        # Filtrar escolas com orçamento vencido
        data_atual = timezone.localdate()
        
        # Verificar se existem escolas com orçamento vencido
        escolas_vencidas = [escola for escola in escolas_query if 
                          escola.data_validade_budget and  # Tem data de validade
                          escola.data_validade_budget < data_atual]  # Data de validade menor que atual
        
        # Filtrar escolas válidas
        escolas = [escola for escola in escolas_query if 
                 not escola.data_validade_budget or  # Sem data de validade (válido para sempre)
                 escola.data_validade_budget >= data_atual]  # Data de validade maior ou igual à atual
        
        # Mostrar mensagem apenas se houver escolas vencidas
        if escolas_vencidas:
            messages.info(request, 'Alguns contratos com orçamento vencido foram ocultados da lista.')
    
    # Ordenar a lista de escolas por nome
    escolas.sort(key=lambda escola: escola.nome)
    
    # Carregar produtos
    produtos = Produto.objects.filter(ativo=True).order_by('nome')
    
    # Preparar dados para o template
    pedido = None  # Novo pedido, então é None
    
    return render(request, 'pedidos/form.html', {
        'pedido': pedido,
        'escolas': escolas,
        'produtos': produtos,
        'status_choices': Pedido.STATUS_CHOICES
    })

@login_required
@permission_required('pedidos.view_pedido')
def detalhes(request, pk):
    """Exibe os detalhes de um pedido"""
    pedido = get_object_or_404(Pedido, pk=pk)
    
    # Verifica se o usuário é um supervisor (Operador de Pedidos)
    is_operador_pedidos = request.user.groups.filter(name='Operador de Pedidos').exists()
    
    if is_operador_pedidos:
        try:
            # Obtém o supervisor associado ao usuário logado
            supervisor = Supervisor.objects.get(usuario=request.user)
            
            # Verifica se a escola do pedido está associada ao supervisor
            if pedido.escola.supervisor != supervisor:
                messages.error(request, 'Você não tem permissão para visualizar os detalhes deste pedido.')
                return redirect('pedidos:lista')
                
        except Supervisor.DoesNotExist:
            # Supervisor não encontrado, redireciona
            messages.warning(request, 'Sua conta de usuário não está corretamente vinculada a um supervisor. Por favor, contate o administrador.')
            return redirect('pedidos:lista')
    
    # O usuário tem permissão para ver o pedido
    logs = pedido.logs.all().select_related('usuario').order_by('-data_hora')
    
    # Verificar o orçamento do contrato
    contexto = {
        'pedido': pedido,
        'logs': logs,
        'pode_ver_logs': not is_operador_pedidos,  # Apenas usuários que não são Operador de Pedidos podem ver logs
        'hoje_iso': timezone.now().date().isoformat()  # Passar a data atual em formato ISO
    }
    
    # Se existe um budget definido, calcular o uso
    if pedido.escola.budget > 0:
        # Verificar se o orçamento está dentro da validade
        if hasattr(pedido.escola, 'data_validade_budget') and pedido.escola.data_validade_budget:
            data_atual = timezone.localdate()
            if pedido.escola.data_validade_budget < data_atual:
                messages.error(
                    request, 
                    "⚠️ ORÇAMENTO EXPIRADO! Este contrato está com o orçamento vencido desde " + 
                    pedido.escola.data_validade_budget.strftime('%d/%m/%Y') + "."
                )
                
                # Adicionar ao contexto
                contexto.update({
                    'budget_contrato': pedido.escola.budget,
                    'budget_expirado': True,
                    'data_validade_budget': pedido.escola.data_validade_budget
                })
                
                return render(request, 'pedidos/detalhes.html', contexto)
        
        # Calcular o valor total de todos os pedidos ativos da escola (excluindo cancelados)
        pedidos_ativos = Pedido.objects.filter(
            escola=pedido.escola, 
            status__in=['pendente', 'aprovado', 'pedido_enviado', 'entregue']
        )
        
        # Calcular o valor total de todos os pedidos manualmente
        valor_total_pedidos = 0
        for p in pedidos_ativos:
            itens_pedido = ItemPedido.objects.filter(pedido=p)
            for item in itens_pedido:
                valor_total_pedidos += item.quantidade * item.valor_unitario
        
        percentual_uso = (valor_total_pedidos / pedido.escola.budget) * 100
        
        # Adicionar ao contexto
        contexto.update({
            'budget_contrato': pedido.escola.budget,
            'valor_total_pedidos': valor_total_pedidos,
            'percentual_uso': percentual_uso
        })
        
        # Exibir mensagem de alerta se estiver próximo ou acima do limite
        if percentual_uso > 100:
            messages.error(
                request, 
                "⚠️ LIMITE DE ORÇAMENTO EXCEDIDO! Este contrato já ultrapassou o valor máximo disponível para pedidos."
            )
        elif percentual_uso >= 90:
            messages.warning(
                request, 
                "⚠️ ATENÇÃO: Este contrato está próximo de atingir o limite máximo do orçamento disponível."
            )
    
    return render(request, 'pedidos/detalhes.html', contexto)

@login_required
@permission_required('pedidos.view_pedido')
def exportar(request):
    """Exporta pedidos para diversos formatos (PDF, Excel, TXT)"""
    # Obter formato de exportação (padrão: excel)
    formato = request.GET.get('formato', 'excel')
    
    # Verificar se há pedidos específicos para exportar
    pedidos_ids = request.GET.get('pedidos_ids')
    
    # Obter pedidos
    if pedidos_ids:
        # Converte a string com IDs separados por vírgula em uma lista
        ids_list = [int(id) for id in pedidos_ids.split(',') if id.isdigit()]
        # Filtra apenas os pedidos cujos IDs estão na lista
        pedidos = Pedido.objects.filter(id__in=ids_list).order_by('-data_solicitacao')
    else:
        # Se não houver IDs específicos, busca todos os pedidos (com filtros opcionais)
        pedidos = Pedido.objects.all().order_by('-data_solicitacao')
        
        # Filtrar se necessário
        status = request.GET.get('status')
        escola_id = request.GET.get('escola')
        produto_id = request.GET.get('produto')
        supervisor_id = request.GET.get('supervisor')
        data_inicio = request.GET.get('data_inicio')
        data_fim = request.GET.get('data_fim')
        
        if status:
            pedidos = pedidos.filter(status=status)
        if escola_id:
            pedidos = pedidos.filter(escola_id=escola_id)
        if produto_id:
            # Filtra pedidos que contêm o produto específico
            pedidos = pedidos.filter(itens__produto_id=produto_id).distinct()
        if supervisor_id:
            # Filtra pedidos de escolas vinculadas a um supervisor específico
            pedidos = pedidos.filter(escola__supervisor_id=supervisor_id)
        if data_inicio:
            pedidos = pedidos.filter(data_solicitacao__gte=data_inicio)
        if data_fim:
            pedidos = pedidos.filter(data_solicitacao__lte=data_fim)
    
    # Verificar se o usuário é um supervisor (Operador de Pedidos)
    if request.user.groups.filter(name='Operador de Pedidos').exists():
        try:
            # Obtém o supervisor associado ao usuário logado
            supervisor = Supervisor.objects.get(usuario=request.user)
            
            # Filtrar pedidos apenas das escolas que o supervisor gerencia
            pedidos = pedidos.filter(escola__supervisor=supervisor)
            
        except Supervisor.DoesNotExist:
            # Supervisor não encontrado, não mostra pedidos
            pedidos = Pedido.objects.none()
    
    if formato == 'excel':
        # Exportar para Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=pedidos.xlsx'
        
        # Criar workbook
        import xlsxwriter
        workbook = xlsxwriter.Workbook(response)
        
        # Formatos
        bold = workbook.add_format({'bold': True})
        title = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center'})
        header = workbook.add_format({'bold': True, 'bg_color': '#D9EAD3', 'border': 1})
        date_format = workbook.add_format({'num_format': 'dd/mm/yyyy hh:mm'})
        money_format = workbook.add_format({'num_format': 'R$ #,##0.00'})
        text_format = workbook.add_format({'text_wrap': True})
        section_title = workbook.add_format({'bold': True, 'font_size': 12, 'bg_color': '#E6E6E6'})
        pedido_header = workbook.add_format({'bold': True, 'bg_color': '#F2F2F2', 'border': 1})
        item_header = workbook.add_format({'bold': True, 'bg_color': '#CFE2F3', 'border': 1})
        cell_border = workbook.add_format({'border': 1})
        detail_title = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center'})
        
        # PLANILHA DETALHADA - único formato mantido conforme solicitado
        detail_sheet = workbook.add_worksheet("Detalhes dos Pedidos")
        
        # Título da planilha detalhada
        detail_sheet.merge_range('A1:G1', 'RELATÓRIO DETALHADO DE PEDIDOS', detail_title)
        detail_sheet.set_row(0, 25)
        
        # Linha atual para inserção de dados
        row = 2
        
        # Para cada pedido, incluir cabeçalho e itens
        for pedido in pedidos:
            # Cabeçalho do pedido
            detail_sheet.merge_range(f'A{row+1}:G{row+1}', f'PEDIDO #{pedido.id}', section_title)
            row += 1
            
            # Informações do pedido
            detail_sheet.write(row, 0, 'Escola:', bold)
            detail_sheet.merge_range(f'B{row+1}:G{row+1}', pedido.escola.nome, workbook.add_format())
            row += 1
            
            detail_sheet.write(row, 0, 'Empresa:', bold)
            detail_sheet.merge_range(f'B{row+1}:G{row+1}', pedido.escola.empresa or '-', workbook.add_format())
            row += 1
            
            detail_sheet.write(row, 0, 'Lote:', bold)
            detail_sheet.merge_range(f'B{row+1}:G{row+1}', pedido.escola.lote or '-', workbook.add_format())
            row += 1
            
            detail_sheet.write(row, 0, 'Data Solicitação:', bold)
            detail_sheet.merge_range(f'B{row+1}:G{row+1}', pedido.data_solicitacao.strftime('%d/%m/%Y %H:%M'), workbook.add_format())
            row += 1
            
            if pedido.data_aprovacao:
                detail_sheet.write(row, 0, 'Data Aprovação:', bold)
                detail_sheet.merge_range(f'B{row+1}:G{row+1}', pedido.data_aprovacao.strftime('%d/%m/%Y %H:%M'), workbook.add_format())
                row += 1
            
            detail_sheet.write(row, 0, 'Status:', bold)
            detail_sheet.merge_range(f'B{row+1}:G{row+1}', pedido.get_status_display(), workbook.add_format())
            row += 1
            
            detail_sheet.write(row, 0, 'Valor Total:', bold)
            detail_sheet.write(row, 1, float(pedido.valor_total), money_format)
            row += 1
            
            detail_sheet.write(row, 0, 'Quantidade de Itens:', bold)
            detail_sheet.write(row, 1, pedido.itens.count())
            row += 2
            
            # Cabeçalho dos itens
            detail_sheet.write(row, 0, 'Produto', item_header)
            detail_sheet.write(row, 1, 'Quantidade', item_header)
            detail_sheet.write(row, 2, 'Valor Unitário', item_header)
            detail_sheet.write(row, 3, 'Valor Total Item', item_header)
            row += 1
            
            # Listar os itens do pedido
            for item in pedido.itens.all():
                detail_sheet.write(row, 0, item.produto.nome, cell_border)
                detail_sheet.write(row, 1, item.quantidade, cell_border)
                detail_sheet.write(row, 2, float(item.valor_unitario), money_format)
                detail_sheet.write(row, 3, float(item.valor_total), money_format)
                row += 1
            
            # Espaço entre pedidos
            row += 2
        
        # Ajustar largura das colunas da planilha de detalhes
        detail_sheet.set_column(0, 0, 30)  # Produto/Título
        detail_sheet.set_column(1, 1, 15)  # Quantidade/Valor
        detail_sheet.set_column(2, 2, 15)  # Valor Unitário
        detail_sheet.set_column(3, 3, 15)  # Valor Total Item
        detail_sheet.set_column(4, 6, 20)  # Outros campos
        
        workbook.close()
        return response
        
    elif formato == 'pdf':
        # Exportar para PDF - versão alternativa sem WeasyPrint
        from django.template.loader import render_to_string
        
        # Como temos problemas de dependências com WeasyPrint no Windows,
        # vamos oferecer uma alternativa de visualização HTML que pode ser impressa como PDF
        response = HttpResponse(content_type='text/html')
        response['Content-Disposition'] = 'inline; filename="pedidos.html"'
        
        # Renderizar template com os dados
        html_string = render_to_string('pedidos/pdf_template.html', 
                                      {'pedidos': pedidos, 'titulo': 'Relatório de Pedidos'})
        
        # Adicionar script para imprimir automaticamente
        html_string += """
        <script>
            window.onload = function() {
                // Adiciona um botão para imprimir
                var btnContainer = document.createElement('div');
                btnContainer.style.textAlign = 'center';
                btnContainer.style.margin = '20px 0';
                
                var printBtn = document.createElement('button');
                printBtn.innerText = 'Imprimir como PDF';
                printBtn.style.padding = '10px 20px';
                printBtn.style.backgroundColor = '#3D5A80';
                printBtn.style.color = 'white';
                printBtn.style.border = 'none';
                printBtn.style.borderRadius = '5px';
                printBtn.style.cursor = 'pointer';
                printBtn.onclick = function() { window.print(); };
                
                btnContainer.appendChild(printBtn);
                document.body.insertBefore(btnContainer, document.body.firstChild);
                
                // Adiciona instruções
                var instructions = document.createElement('p');
                instructions.innerText = 'Para salvar como PDF, use a opção "Salvar como PDF" na janela de impressão.';
                instructions.style.textAlign = 'center';
                instructions.style.color = '#666';
                btnContainer.appendChild(instructions);
            };
        </script>
        <style>
            @media print {
                button, p {
                    display: none;
                }
            }
        </style>
        """
        
        response.write(html_string)
        return response
        
    elif formato == 'txt':
        # Exportar para TXT
        response = HttpResponse(content_type='text/plain')
        response['Content-Disposition'] = 'attachment; filename="pedidos.txt"'
        
        # Cabeçalho
        response.write("RELATÓRIO DE PEDIDOS\n")
        response.write("=" * 80 + "\n\n")
        
        # Dados
        total_geral = 0
        for pedido in pedidos:
            total_geral += pedido.valor_total
            response.write(f"PEDIDO #{pedido.id}\n")
            response.write("-" * 40 + "\n")
            response.write(f"Escola: {pedido.escola.nome}\n")
            response.write(f"Empresa: {pedido.escola.empresa or '-'}\n")
            response.write(f"Lote: {pedido.escola.lote or '-'}\n")
            response.write(f"Data de Solicitação: {pedido.data_solicitacao.strftime('%d/%m/%Y %H:%M')}\n")
            
            if pedido.data_aprovacao:
                response.write(f"Data de Aprovação: {pedido.data_aprovacao.strftime('%d/%m/%Y %H:%M')}\n")
            
            response.write(f"Status: {pedido.get_status_display()}\n")
            response.write(f"Valor Total: R$ {pedido.valor_total:.2f}\n")
            response.write(f"Quantidade de Itens: {pedido.itens.count()}\n\n")
            
            # Detalhes dos itens
            response.write("ITENS DO PEDIDO:\n")
            response.write("-" * 40 + "\n")
            response.write(f"{'Produto':<40} {'Qtd':<8} {'Valor Unit.':<15} {'Valor Total':<15}\n")
            response.write("-" * 80 + "\n")
            
            for item in pedido.itens.all():
                produto_nome = item.produto.nome
                if len(produto_nome) > 37:
                    produto_nome = produto_nome[:37] + "..."
                
                response.write(f"{produto_nome:<40} {item.quantidade:<8} R$ {item.valor_unitario:<10.2f} R$ {item.valor_total:<10.2f}\n")
            
            response.write("\n" + "=" * 80 + "\n\n")
        
        # Resumo final
        response.write(f"\nTotal de Pedidos: {pedidos.count()}\n")
        response.write(f"Valor Total Geral: R$ {total_geral:.2f}\n")
        response.write("\n\nFim do Relatório\n")
        
        return response
    
    else:
        return HttpResponse("Formato não suportado")

@login_required
@permission_required('pedidos.view_pedido')
def relatorio(request):
    """Gera relatório de pedidos"""
    # Obter parâmetros do filtro
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    escola_id = request.GET.get('escola')
    empresa = request.GET.get('empresa')
    status = request.GET.get('status')
    
    # Iniciar conjunto de pedidos
    pedidos = Pedido.objects.all()
    
    # Verificar se o usuário é um supervisor (Operador de Pedidos)
    is_supervisor = False
    supervisor = None
    supervisor_escolas_ids = []
    
    if request.user.groups.filter(name='Operador de Pedidos').exists():
        is_supervisor = True
        try:
            # Obtém o supervisor associado ao usuário logado
            supervisor = Supervisor.objects.get(usuario=request.user)
            
            # Filtrar pedidos apenas das escolas que o supervisor gerencia
            pedidos = pedidos.filter(escola__supervisor=supervisor)
            
            # Obter IDs das escolas para filtros
            supervisor_escolas_ids = list(Escola.objects.filter(supervisor=supervisor, ativo=True).values_list('id', flat=True))
            
            # Mensagem informativa
            messages.info(request, f'Você está visualizando relatórios apenas de contratos sob sua supervisão.')
            
        except Supervisor.DoesNotExist:
            # Supervisor não encontrado, não mostra pedidos
            pedidos = Pedido.objects.none()
            messages.warning(request, 'Sua conta de usuário não está corretamente vinculada a um supervisor. Por favor, contate o administrador.')
    
    # Aplicar filtros adicionais
    if data_inicio:
        pedidos = pedidos.filter(data_solicitacao__gte=data_inicio)
    if data_fim:
        pedidos = pedidos.filter(data_solicitacao__lte=data_fim)
    if escola_id:
        # Verificar se o supervisor pode acessar essa escola
        if is_supervisor and int(escola_id) not in supervisor_escolas_ids:
            messages.warning(request, 'Você não tem permissão para visualizar os pedidos desta escola.')
            escola_id = None
        else:
            pedidos = pedidos.filter(escola_id=escola_id)
    if empresa:
        pedidos = pedidos.filter(escola__empresa=empresa)
    if status:
        pedidos = pedidos.filter(status=status)
    
    # Preparar estatísticas
    total_pedidos = pedidos.count()
    valor_total = sum(pedido.valor_total for pedido in pedidos)
    total_escolas = pedidos.values('escola').distinct().count()
    ticket_medio = valor_total / total_pedidos if total_pedidos > 0 else 0
    
    # Contagem por status
    status_pendente = pedidos.filter(status='pendente').count()
    status_aprovado = pedidos.filter(status='aprovado').count()
    status_pedido_enviado = pedidos.filter(status='pedido_enviado').count()
    status_entregue = pedidos.filter(status='entregue').count()
    status_cancelado = pedidos.filter(status='cancelado').count()
    
    # Dados para o gráfico de meses (simplificado)
    meses_labels = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    meses_valores = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]  # Valores reais seriam calculados a partir dos dados
    
    # Carregar escolas para o filtro
    if is_supervisor and supervisor:
        escolas = Escola.objects.filter(supervisor=supervisor, ativo=True).order_by('nome')
        # Obter lista de empresas únicas das escolas do supervisor
        empresas = escolas.exclude(empresa='').exclude(empresa__isnull=True).values('empresa').distinct().order_by('empresa')
    else:
        escolas = Escola.objects.filter(ativo=True).order_by('nome')
        # Obter lista de empresas únicas
        empresas = escolas.exclude(empresa='').exclude(empresa__isnull=True).values('empresa').distinct().order_by('empresa')
    
    # Preparar filtros para passar de volta ao template
    filtros = {
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'escola': escola_id,
        'empresa': empresa,
        'status': status
    }
    
    # Preparar estatísticas
    estatisticas = {
        'total_pedidos': total_pedidos,
        'valor_total': valor_total,
        'total_escolas': total_escolas,
        'ticket_medio': ticket_medio,
        'status_pendente': status_pendente,
        'status_aprovado': status_aprovado,
        'status_pedido_enviado': status_pedido_enviado,
        'status_entregue': status_entregue,
        'status_cancelado': status_cancelado,
        'meses_labels': meses_labels,
        'meses_valores': meses_valores
    }
    
    return render(request, 'pedidos/relatorio.html', {
        'pedidos': pedidos,
        'escolas': escolas,
        'empresas': empresas,
        'filtros': filtros,
        'estatisticas': estatisticas
    })

@login_required
@permission_required('pedidos.view_pedido')
def dashboard_relatorios(request):
    """Dashboard central de relatórios com análises e estatísticas"""
    
    # Período de análise (padrão: últimos 12 meses)
    periodo = request.GET.get('periodo', '12m')
    
    hoje = timezone.now().date()
    
    if periodo == '30d':
        data_inicio = hoje - timedelta(days=30)
        titulo_periodo = "Últimos 30 dias"
    elif periodo == '90d':
        data_inicio = hoje - timedelta(days=90)
        titulo_periodo = "Últimos 90 dias"
    elif periodo == '6m':
        data_inicio = hoje - timedelta(days=180)
        titulo_periodo = "Últimos 6 meses"
    elif periodo == 'ano':
        data_inicio = datetime(hoje.year, 1, 1).date()
        titulo_periodo = f"Ano atual ({hoje.year})"
    else:  # 12m (padrão)
        data_inicio = hoje - timedelta(days=365)
        titulo_periodo = "Últimos 12 meses"
    
    # Filtros adicionais (opcionais)
    empresa = request.GET.get('empresa')
    lote = request.GET.get('lote')
    
    # -------------- ANÁLISE DE PEDIDOS --------------
    # Base de pedidos do período
    pedidos_periodo = Pedido.objects.filter(data_solicitacao__date__gte=data_inicio)
    
    # Verificar se o usuário é um supervisor (Operador de Pedidos)
    is_supervisor = False
    supervisor_escolas = None
    
    if request.user.groups.filter(name='Operador de Pedidos').exists():
        is_supervisor = True
        try:
            # Obtém o supervisor associado ao usuário logado
            supervisor = Supervisor.objects.get(usuario=request.user)
            
            # Filtrar pedidos apenas das escolas que o supervisor gerencia
            pedidos_periodo = pedidos_periodo.filter(escola__supervisor=supervisor)
            
            # Obter escolas do supervisor
            supervisor_escolas = Escola.objects.filter(supervisor=supervisor, ativo=True)
            
            # Mensagem informativa
            messages.info(request, f'Você está visualizando estatísticas apenas de contratos sob sua supervisão.')
            
        except Supervisor.DoesNotExist:
            # Supervisor não encontrado, não mostra pedidos
            pedidos_periodo = Pedido.objects.none()
            messages.warning(request, 'Sua conta de usuário não está corretamente vinculada a um supervisor. Por favor, contate o administrador.')
    
    # Aplicar filtros adicionais
    if empresa:
        pedidos_periodo = pedidos_periodo.filter(escola__empresa=empresa)
    if lote:
        pedidos_periodo = pedidos_periodo.filter(escola__lote=lote)
    
    # Estatísticas gerais
    total_pedidos = pedidos_periodo.count()
    valor_total_periodo = sum(pedido.valor_total for pedido in pedidos_periodo)
    ticket_medio = valor_total_periodo / total_pedidos if total_pedidos > 0 else 0
    qtd_itens_total = sum(pedido.itens.count() for pedido in pedidos_periodo)
    media_itens_por_pedido = qtd_itens_total / total_pedidos if total_pedidos > 0 else 0
    
    # Pedidos por status
    pedidos_por_status = {
        'pendente': pedidos_periodo.filter(status='pendente').count(),
        'aprovado': pedidos_periodo.filter(status='aprovado').count(),
        'pedido_enviado': pedidos_periodo.filter(status='pedido_enviado').count(),
        'entregue': pedidos_periodo.filter(status='entregue').count(),
        'cancelado': pedidos_periodo.filter(status='cancelado').count(),
    }
    
    # Análise temporal - Pedidos por mês (últimos 12 meses)
    pedidos_por_mes = []
    for i in range(12):
        mes_data = hoje.replace(day=1) - timedelta(days=i*30)
        mes_inicio = mes_data.replace(day=1)
        
        # Cálculo seguro do último dia do mês
        if mes_data.month == 12:
            proximo_mes = datetime(mes_data.year+1, 1, 1)
        else:
            proximo_mes = datetime(mes_data.year, mes_data.month+1, 1)
        
        mes_fim = proximo_mes - timedelta(days=1)
        
        qtd_pedidos = pedidos_periodo.filter(
            data_solicitacao__date__gte=mes_inicio,
            data_solicitacao__date__lte=mes_fim
        ).count()
        
        valor_pedidos = sum(p.valor_total for p in pedidos_periodo.filter(
            data_solicitacao__date__gte=mes_inicio,
            data_solicitacao__date__lte=mes_fim
        ))
        
        pedidos_por_mes.append({
            'mes': mes_data.strftime('%b/%Y'),
            'qtd': qtd_pedidos,
            'valor': round(float(valor_pedidos), 2)
        })
    
    # Invertendo para ordem cronológica
    pedidos_por_mes.reverse()
    
    # -------------- ANÁLISE DE ESCOLAS --------------
    # Decidir quais escolas incluir na análise
    if is_supervisor and supervisor_escolas is not None:
        escolas_para_analise = supervisor_escolas
    else:
        escolas_para_analise = Escola.objects.filter(ativo=True)
    
    # Top escolas por volume de pedidos
    top_escolas_pedidos = escolas_para_analise.filter(
        pedidos__data_solicitacao__date__gte=data_inicio
    ).annotate(
        qtd_pedidos=Count('pedidos')
    ).order_by('-qtd_pedidos')[:10]
    
    # Top escolas por valor de pedidos
    top_escolas_valor = []
    escolas_com_pedidos = escolas_para_analise.filter(pedidos__data_solicitacao__date__gte=data_inicio).distinct()
    for escola in escolas_com_pedidos:
        valor_total = sum(p.valor_total for p in escola.pedidos.filter(data_solicitacao__date__gte=data_inicio))
        if valor_total > 0:
            top_escolas_valor.append({
                'escola': escola,
                'valor_total': valor_total
            })
    
    # Ordenar por valor total e limitar a 10
    top_escolas_valor = sorted(top_escolas_valor, key=lambda x: x['valor_total'], reverse=True)[:10]
    
    # -------------- ANÁLISE DE PRODUTOS --------------
    # Produtos mais pedidos
    produtos_mais_pedidos = []
    produtos_dict = {}
    
    # Itens de pedidos no período selecionado
    itens_periodo = ItemPedido.objects.filter(pedido__data_solicitacao__date__gte=data_inicio)
    if is_supervisor and supervisor_escolas is not None:
        itens_periodo = itens_periodo.filter(pedido__escola__in=supervisor_escolas)
    if empresa:
        itens_periodo = itens_periodo.filter(pedido__escola__empresa=empresa)
    if lote:
        itens_periodo = itens_periodo.filter(pedido__escola__lote=lote)
    
    # Agrupar quantidades por produto
    for item in itens_periodo:
        produto_id = item.produto.id
        if produto_id in produtos_dict:
            produtos_dict[produto_id]['quantidade'] += item.quantidade
            produtos_dict[produto_id]['valor_total'] += item.valor_total
        else:
            produtos_dict[produto_id] = {
                'produto': item.produto,
                'quantidade': item.quantidade,
                'valor_total': item.valor_total
            }
    
    # Converter para lista e ordenar
    produtos_mais_pedidos = list(produtos_dict.values())
    top_produtos_qtd = sorted(produtos_mais_pedidos, key=lambda x: x['quantidade'], reverse=True)[:10]
    top_produtos_valor = sorted(produtos_mais_pedidos, key=lambda x: x['valor_total'], reverse=True)[:10]
    
    # -------------- ANÁLISE DE EMPRESAS E LOTES --------------
    # Distribuição por empresa
    pedidos_por_empresa = pedidos_periodo.values('escola__empresa').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Distribuição por lote
    pedidos_por_lote = pedidos_periodo.values('escola__lote').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Lista de empresas e lotes para filtros
    empresas = Escola.objects.exclude(empresa='').exclude(empresa__isnull=True).values_list('empresa', flat=True).distinct().order_by('empresa')
    lotes = Escola.objects.exclude(lote='').exclude(lote__isnull=True).values_list('lote', flat=True).distinct().order_by('lote')
    
    # Lista de produtos ativos para os filtros
    produtos = Produto.objects.filter(ativo=True).order_by('nome')
    
    # Lista de escolas ativas para os filtros
    escolas = Escola.objects.filter(ativo=True).order_by('nome')
    
    # Lista de supervisores ativos para os filtros
    supervisores = Supervisor.objects.filter(ativo=True).order_by('nome')
    
    # Preparar o contexto para o template
    context = {
        'titulo_periodo': titulo_periodo,
        'periodo': periodo,
        'empresa_filtro': empresa,
        'lote_filtro': lote,
        'total_pedidos': total_pedidos,
        'valor_total_periodo': valor_total_periodo,
        'ticket_medio': ticket_medio,
        'media_itens_por_pedido': media_itens_por_pedido,
        'pedidos_por_status': pedidos_por_status,
        'pedidos_por_mes': pedidos_por_mes,
        'top_escolas_pedidos': top_escolas_pedidos,
        'top_escolas_valor': top_escolas_valor,
        'top_produtos_qtd': top_produtos_qtd,
        'top_produtos_valor': top_produtos_valor,
        'pedidos_por_empresa': pedidos_por_empresa,
        'pedidos_por_lote': pedidos_por_lote,
        'empresas': empresas,
        'lotes': lotes,
        'produtos': produtos,
        'escolas': escolas,
        'supervisores': supervisores
    }
    
    return render(request, 'pedidos/dashboard_relatorios.html', context)

@login_required
@permission_required('pedidos.delete_pedido')
def excluir(request, pk):
    """Exclui um pedido caso ele não esteja aprovado"""
    pedido = get_object_or_404(Pedido, pk=pk)
    
    # Verificar se o usuário tem permissão para excluir este pedido específico
    pode_excluir = True
    
    # Verificação para supervisores
    if request.user.groups.filter(name='Operador de Pedidos').exists():
        try:
            supervisor = Supervisor.objects.get(usuario=request.user)
            
            # Verifica se o supervisor gerencia a escola do pedido
            if pedido.escola.supervisor != supervisor:
                messages.error(request, 'Você não tem permissão para excluir este pedido.')
                return redirect('pedidos:lista')
                
            # Operadores de Pedido só podem excluir pedidos não aprovados
            if pedido.status in ['aprovado', 'pedido_enviado', 'entregue']:
                pode_excluir = False
                messages.error(request, 'Como Operador de Pedidos, você só pode excluir pedidos que não foram aprovados.')
                return redirect('pedidos:detalhes', pk=pedido.pk)
                
        except Supervisor.DoesNotExist:
            messages.warning(request, 'Sua conta de usuário não está corretamente vinculada a um supervisor. Por favor, contate o administrador.')
            return redirect('pedidos:lista')
    
    # Para outros perfis, verificar o status do pedido (regra de negócio geral)
    if pedido.status in ['aprovado', 'pedido_enviado', 'entregue']:
        messages.error(request, 'Este pedido não pode ser excluído pois já foi aprovado ou processado.')
        return redirect('pedidos:detalhes', pk=pedido.pk)
    
    # Se chegou aqui, o pedido pode ser excluído
    nome = f"Pedido #{pedido.pk}"
    pedido_id = pedido.pk
    escola_nome = pedido.escola.nome
    
    # Criar um registro de log antes de excluir o pedido
    # Como o pedido será excluído, não podemos relacionar este log diretamente
    # Vamos armazenar a informação em um log separado
    
    # Registrar log de exclusão
    descricao = f"Pedido #{pedido_id} da escola {escola_nome} foi excluído por {request.user.get_full_name() or request.user.username}"
    
    # Cria um log de sistema
    try:
        from admin_logs.models import SystemLog
        SystemLog.objects.create(
            usuario=request.user,
            acao="exclusao_pedido",
            detalhes=descricao
        )
    except ImportError:
        # Caso não exista o app de logs do sistema, apenas seguimos
        pass
    
    # Exclui o pedido
    pedido.delete()
    
    messages.success(request, f'{nome} foi excluído com sucesso!')
    return redirect('pedidos:lista')

@login_required
@permission_required('pedidos.add_pedido')
def duplicar(request, pk):
    """Duplica um pedido existente, criando uma nova versão"""
    pedido_original = get_object_or_404(Pedido, pk=pk)
    
    # Verifica se o usuário é um supervisor (Operador de Pedidos)
    if request.user.groups.filter(name='Operador de Pedidos').exists():
        try:
            # Obtém o supervisor associado ao usuário logado
            supervisor = Supervisor.objects.get(usuario=request.user)
            
            # Verifica se a escola do pedido está associada ao supervisor
            if pedido_original.escola.supervisor != supervisor:
                messages.error(request, 'Você não tem permissão para duplicar este pedido.')
                return redirect('pedidos:lista')
                
        except Supervisor.DoesNotExist:
            # Supervisor não encontrado, redireciona
            messages.warning(request, 'Sua conta de usuário não está corretamente vinculada a um supervisor. Por favor, contate o administrador.')
            return redirect('pedidos:lista')
    
    try:
        with transaction.atomic():
            # Criar um novo pedido baseado no original
            novo_pedido = Pedido(
                escola=pedido_original.escola,
                data_solicitacao=timezone.now(),  # Data atual para o novo pedido
                observacoes=f"Duplicado do Pedido #{pedido_original.id}. {pedido_original.observacoes or ''}",
                status='pendente'  # Sempre começa como pendente
            )
            novo_pedido.save()
            
            # Duplicar itens do pedido
            for item_original in pedido_original.itens.all():
                novo_item = ItemPedido(
                    pedido=novo_pedido,
                    produto=item_original.produto,
                    quantidade=item_original.quantidade,
                    valor_unitario=item_original.produto.valor_unitario  # Usar valor atual do produto
                )
                novo_item.save()
            
            messages.success(request, f"Pedido duplicado com sucesso! Novo número de pedido: #{novo_pedido.id}")
            return redirect('pedidos:detalhes', pk=novo_pedido.pk)
            
    except Exception as e:
        messages.error(request, f"Erro ao duplicar o pedido: {str(e)}")
        return redirect('pedidos:detalhes', pk=pk)

@login_required
@permission_required('pedidos.view_pedido')
def imprimir_pedido(request, pk):
    """Exibe uma versão formatada de um pedido para impressão"""
    pedido = get_object_or_404(Pedido, pk=pk)
    
    # Verifica se o usuário é um supervisor (Operador de Pedidos)
    if request.user.groups.filter(name='Operador de Pedidos').exists():
        try:
            # Obtém o supervisor associado ao usuário logado
            supervisor = Supervisor.objects.get(usuario=request.user)
            
            # Verifica se a escola do pedido está associada ao supervisor
            if pedido.escola.supervisor != supervisor:
                messages.error(request, 'Você não tem permissão para imprimir os detalhes deste pedido.')
                return redirect('pedidos:lista')
                
        except Supervisor.DoesNotExist:
            # Supervisor não encontrado, redireciona
            messages.warning(request, 'Sua conta de usuário não está corretamente vinculada a um supervisor. Por favor, contate o administrador.')
            return redirect('pedidos:lista')
    
    # Buscar logs do pedido para histórico
    logs = pedido.logs.all().select_related('usuario').order_by('-data_hora')
    
    # O logo está na pasta de imagens estáticas
    logo_path = f"{settings.STATIC_URL}images/logo.png"
    
    return render(request, 'pedidos/imprimir_pedido.html', {
        'pedido': pedido,
        'logs': logs,
        'data_hoje': timezone.now(),
        'logo_path': logo_path,
    })

@login_required
@permission_required('pedidos.view_pedido')
def central_relatorios(request):
    """Central de criação e gestão de relatórios personalizados"""
    return render(request, 'pedidos/central_relatorios.html')

@login_required
@permission_required('pedidos.view_pedido')
def exportar_excel(request):
    """Exporta dados de pedidos para Excel com formatação personalizada"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import json
    from datetime import datetime

    # Criar um novo workbook e selecionar a planilha ativa
    wb = Workbook()
    ws = wb.active
    
    # Obter parâmetros do formulário
    titulo = request.POST.get('report_title', 'Relatório de Pedidos')
    incluir_filtros = request.POST.get('include_filters') == '1'
    incluir_resumo = request.POST.get('include_summary') == '1'
    incluir_graficos = request.POST.get('include_charts') == '1'
    
    # Estilo do cabeçalho
    estilo_cabecalho = request.POST.get('header_style', 'default')
    fonte_familia = request.POST.get('font_family', 'calibri')
    
    # Opções de formatação
    alternar_cores = request.POST.get('alternate_row_colors') == '1'
    add_filtros_auto = request.POST.get('add_auto_filters') == '1'
    congelar_cabecalho = request.POST.get('freeze_header_row') == '1'
    
    # Filtros de dados
    periodo = request.POST.get('period', 'all')
    status_selecionados = request.POST.getlist('status')
    
    # Configurar período de datas
    hoje = timezone.now().date()
    data_inicio = None
    data_fim = hoje
    
    if periodo == 'current_month':
        data_inicio = datetime(hoje.year, hoje.month, 1).date()
    elif periodo == 'last_month':
        if hoje.month == 1:
            data_inicio = datetime(hoje.year - 1, 12, 1).date()
            data_fim = datetime(hoje.year, 1, 1).date() - timedelta(days=1)
        else:
            data_inicio = datetime(hoje.year, hoje.month - 1, 1).date()
            data_fim = datetime(hoje.year, hoje.month, 1).date() - timedelta(days=1)
    elif periodo == 'last_90d':
        data_inicio = hoje - timedelta(days=90)
    elif periodo == 'custom':
        try:
            data_inicio = datetime.strptime(request.POST.get('start_date'), '%Y-%m-%d').date()
            data_fim = datetime.strptime(request.POST.get('end_date'), '%Y-%m-%d').date()
        except (ValueError, TypeError):
            data_inicio = hoje - timedelta(days=365)  # Fallback para 1 ano
    else:  # 'all' ou fallback
        data_inicio = hoje - timedelta(days=365 * 5)  # Últimos 5 anos como "todos"
    
    # Consultar pedidos com base nos filtros
    pedidos = Pedido.objects.all()
    
    # Verificar se o usuário é um supervisor (Operador de Pedidos)
    if request.user.groups.filter(name='Operador de Pedidos').exists():
        try:
            # Obtém o supervisor associado ao usuário logado
            supervisor = Supervisor.objects.get(usuario=request.user)
            
            # Filtrar pedidos apenas das escolas que o supervisor gerencia
            pedidos = pedidos.filter(escola__supervisor=supervisor)
            
        except Supervisor.DoesNotExist:
            # Supervisor não encontrado, não mostra pedidos
            pedidos = Pedido.objects.none()
    
    if data_inicio:
        pedidos = pedidos.filter(data_solicitacao__date__gte=data_inicio)
    if data_fim:
        pedidos = pedidos.filter(data_solicitacao__date__lte=data_fim)
    
    if status_selecionados:
        pedidos = pedidos.filter(status__in=status_selecionados)
    
    # Configurações de estilo com base nas preferências
    # Cores para cabeçalhos
    cores_cabecalho = {
        'default': "0070C0",  # Azul
        'green': "00B050",    # Verde
        'red': "C00000",      # Vermelho
        'purple': "7030A0",   # Roxo
        'gray': "808080"      # Cinza
    }
    
    # Fontes
    fontes = {
        'calibri': "Calibri",
        'arial': "Arial",
        'times_new_roman': "Times New Roman",
        'century_gothic': "Century Gothic"
    }
    
    # Configuração de cores e estilos
    cor_cabecalho = cores_cabecalho.get(estilo_cabecalho, cores_cabecalho['default'])
    fonte = fontes.get(fonte_familia, fontes['calibri'])
    
    # Estilo para o título do relatório
    titulo_font = Font(name=fonte, size=16, bold=True, color="333333")
    titulo_alignment = Alignment(horizontal='center', vertical='center')
    
    # Estilo para cabeçalhos
    header_font = Font(name=fonte, size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=cor_cabecalho, end_color=cor_cabecalho, fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Estilo para células normais
    normal_font = Font(name=fonte, size=11)
    normal_alignment = Alignment(vertical='center')
    
    # Estilo para linhas alternadas
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Bordas
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # ========== Título do Relatório ==========
    ws.title = "Relatório de Pedidos"
    ws.merge_cells('A1:H1')  # Substituído merge_range por merge_cells
    ws['A1'] = titulo  # Definir o valor da célula A1 após a mesclagem
    ws['A1'].font = titulo_font
    ws['A1'].alignment = titulo_alignment
    ws.row_dimensions[1].height = 30
    
    # ========== Informações do Relatório ==========
    linha_atual = 2
    
    if incluir_filtros:
        ws['A2'] = "Informações do Relatório:"
        ws['A2'].font = Font(name=fonte, size=12, bold=True)
        linha_atual += 1
        
        ws[f'A{linha_atual}'] = "Período:"
        ws[f'B{linha_atual}'] = f"{data_inicio.strftime('%d/%m/%Y')} até {data_fim.strftime('%d/%m/%Y')}"
        
        linha_atual += 1
        ws[f'A{linha_atual}'] = "Filtros:"
        filtros_texto = "Todos os pedidos"
        if status_selecionados:
            filtros_texto = f"Status: {', '.join(status_selecionados)}"
        ws[f'B{linha_atual}'] = filtros_texto
        
        linha_atual += 1
        ws[f'A{linha_atual}'] = "Total de Pedidos:"
        ws[f'B{linha_atual}'] = pedidos.count()
        
        linha_atual += 2  # Espaço em branco
    
    # ========== Resumo ==========
    if incluir_resumo:
        ws[f'A{linha_atual}'] = "Resumo"
        ws[f'A{linha_atual}'].font = Font(name=fonte, size=14, bold=True)
        linha_atual += 1
        
        # Agrupar pedidos por status
        pedidos_por_status = {
            'pendente': pedidos.filter(status='pendente').count(),
            'aprovado': pedidos.filter(status='aprovado').count(),
            'pedido_enviado': pedidos.filter(status='pedido_enviado').count(),
            'entregue': pedidos.filter(status='entregue').count(),
            'cancelado': pedidos.filter(status='cancelado').count(),
        }
        
        # Cabeçalhos do resumo
        cabecalhos_resumo = ['Status', 'Quantidade', 'Percentual']
        for i, header in enumerate(cabecalhos_resumo):
            col = get_column_letter(i + 1)
            ws[f'{col}{linha_atual}'] = header
            ws[f'{col}{linha_atual}'].font = header_font
            ws[f'{col}{linha_atual}'].fill = header_fill
            ws[f'{col}{linha_atual}'].alignment = header_alignment
            ws[f'{col}{linha_atual}'].border = thin_border
        
        linha_atual += 1
        total_pedidos = pedidos.count()
        
        # Dados do resumo
        for status, qtd in pedidos_por_status.items():
            percentual = qtd / total_pedidos * 100 if total_pedidos > 0 else 0
            
            ws[f'A{linha_atual}'] = status.capitalize()
            ws[f'B{linha_atual}'] = qtd
            ws[f'C{linha_atual}'] = f"{percentual:.1f}%"
            
            for col in ['A', 'B', 'C']:
                ws[f'{col}{linha_atual}'].font = normal_font
                ws[f'{col}{linha_atual}'].alignment = normal_alignment
                ws[f'{col}{linha_atual}'].border = thin_border
            
            linha_atual += 1
        
        # Total
        ws[f'A{linha_atual}'] = "Total"
        ws[f'A{linha_atual}'].font = Font(name=fonte, size=11, bold=True)
        ws[f'B{linha_atual}'] = total_pedidos
        ws[f'B{linha_atual}'].font = Font(name=fonte, size=11, bold=True)
        ws[f'C{linha_atual}'] = "100%"
        ws[f'C{linha_atual}'].font = Font(name=fonte, size=11, bold=True)
        
        linha_atual += 2  # Espaço em branco
    
    # ========== Listagem Completa de Pedidos ==========
    ws[f'A{linha_atual}'] = "Listagem de Pedidos"
    ws[f'A{linha_atual}'].font = Font(name=fonte, size=14, bold=True)
    linha_atual += 1
    
    # Cabeçalhos da tabela
    cabecalhos = ['ID', 'Data', 'Escola', 'Endereço', 'Empresa', 'Status', 'Total de Itens', 'Valor Total', 'Itens do Pedido']
    for i, header in enumerate(cabecalhos):
        col = get_column_letter(i + 1)
        ws[f'{col}{linha_atual}'] = header
        ws[f'{col}{linha_atual}'].font = header_font
        ws[f'{col}{linha_atual}'].fill = header_fill
        ws[f'{col}{linha_atual}'].alignment = header_alignment
        ws[f'{col}{linha_atual}'].border = thin_border
    
    # Congelar painel no cabeçalho se solicitado
    if congelar_cabecalho:
        ws.freeze_panes = f'A{linha_atual + 1}'
    
    # Adicionar filtros automáticos se solicitado
    if add_filtros_auto:
        ws.auto_filter.ref = f'A{linha_atual}:{get_column_letter(len(cabecalhos))}{linha_atual}'
    
    linha_atual += 1
    linha_inicial_dados = linha_atual  # Para ajustar larguras de coluna depois
    
    # Preencher com os dados dos pedidos
    for i, pedido in enumerate(pedidos):
        # Obter os itens do pedido
        itens_pedido = pedido.itens.all().select_related('produto')
        itens_texto = ""
        
        # Formatar os itens do pedido como texto
        for item in itens_pedido:
            itens_texto += f"• {item.quantidade}x {item.produto.nome} (R$ {item.valor_unitario:.2f} cada) - Total: R$ {item.valor_total:.2f}\n"
        
        # Dados básicos do pedido
        ws[f'A{linha_atual}'] = pedido.id
        ws[f'B{linha_atual}'] = pedido.data_solicitacao.strftime('%d/%m/%Y')
        ws[f'C{linha_atual}'] = pedido.escola.nome if pedido.escola else 'N/A'
        
        # Adicionar endereço da escola
        endereco_completo = ""
        if pedido.escola:
            escola = pedido.escola
            endereco_partes = []
            
            if escola.endereco:
                endereco_partes.append(escola.endereco)
            if escola.cidade:
                endereco_partes.append(escola.cidade)
            if escola.estado:
                endereco_partes.append(escola.estado)
            if escola.cep:
                endereco_partes.append(f"CEP: {escola.cep}")
            
            endereco_completo = ", ".join(filter(None, endereco_partes))
        
        ws[f'D{linha_atual}'] = endereco_completo or "Endereço não informado"
        ws[f'E{linha_atual}'] = pedido.escola.empresa if pedido.escola else 'N/A'
        ws[f'F{linha_atual}'] = pedido.get_status_display()
        ws[f'G{linha_atual}'] = pedido.itens.count()
        ws[f'H{linha_atual}'] = pedido.valor_total
        ws[f'I{linha_atual}'] = itens_texto
        
        # Aplicar estilo
        for col in range(1, len(cabecalhos) + 1):
            col_letter = get_column_letter(col)
            ws[f'{col_letter}{linha_atual}'].font = normal_font
            ws[f'{col_letter}{linha_atual}'].alignment = normal_alignment
            # Ajustar alinhamento e quebra de texto para a coluna de itens
            if col == 9:  # Coluna de Itens do Pedido
                ws[f'{col_letter}{linha_atual}'].alignment = Alignment(vertical='top', wrap_text=True)
            
            ws[f'{col_letter}{linha_atual}'].border = thin_border
            
            # Formatar valores numéricos
            if col == 8:  # Coluna Valor Total
                ws[f'{col_letter}{linha_atual}'].number_format = 'R$ #,##0.00'
        
        
        # Aplicar cor de fundo em linhas alternadas
        if alternar_cores and i % 2 == 1:
            for col in range(1, len(cabecalhos) + 1):
                col_letter = get_column_letter(col)
                ws[f'{col_letter}{linha_atual}'].fill = alternate_fill
        
        # Altura da linha ajustada para acomodar os itens
        ws.row_dimensions[linha_atual].height = max(20, 15 * (itens_pedido.count() + 1))
        
        linha_atual += 1
    
    # Ajustar larguras das colunas
    for col in range(1, len(cabecalhos) + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 16  # Largura padrão
    
    # Ajustes específicos
    ws.column_dimensions['A'].width = 10   # ID
    ws.column_dimensions['B'].width = 12   # Data
    ws.column_dimensions['C'].width = 30   # Escola
    ws.column_dimensions['D'].width = 40   # Endereço
    ws.column_dimensions['E'].width = 20   # Empresa
    ws.column_dimensions['H'].width = 15   # Valor Total
    ws.column_dimensions['I'].width = 60   # Itens do Pedido
    
    # ========== Criar a segunda planilha com Detalhes de Itens ==========
    ws_itens = wb.create_sheet(title="Detalhes dos Itens")
    
    # Título
    ws_itens.merge_cells('A1:H1')  # Substituído merge_range por merge_cells
    ws_itens['A1'] = f"{titulo} - Detalhes dos Itens"  # Definir o valor na primeira célula
    ws_itens['A1'].font = titulo_font
    ws_itens['A1'].alignment = titulo_alignment
    ws_itens.row_dimensions[1].height = 30
    
    # Cabeçalhos da tabela de itens
    linha_atual = 3
    cabecalhos_itens = ['ID Pedido', 'Produto', 'Código', 'Quantidade', 'Valor Unitário', 'Valor Total', 'Status do Pedido']
    
    for i, header in enumerate(cabecalhos_itens):
        col = get_column_letter(i + 1)
        ws_itens[f'{col}{linha_atual}'] = header
        ws_itens[f'{col}{linha_atual}'].font = header_font
        ws_itens[f'{col}{linha_atual}'].fill = header_fill
        ws_itens[f'{col}{linha_atual}'].alignment = header_alignment
        ws_itens[f'{col}{linha_atual}'].border = thin_border
    
    # Congelar painel no cabeçalho se solicitado
    if congelar_cabecalho:
        ws_itens.freeze_panes = f'A{linha_atual + 1}'
    
    # Adicionar filtros automáticos se solicitado
    if add_filtros_auto:
        ws_itens.auto_filter.ref = f'A{linha_atual}:{get_column_letter(len(cabecalhos_itens))}{linha_atual}'
    
    linha_atual += 1
    
    # Preencher com os dados dos itens de todos os pedidos
    itens = ItemPedido.objects.filter(pedido__in=pedidos).select_related('pedido', 'produto')
    
    for i, item in enumerate(itens):
        ws_itens[f'A{linha_atual}'] = item.pedido.id
        ws_itens[f'B{linha_atual}'] = item.produto.nome
        ws_itens[f'C{linha_atual}'] = item.produto.codigo
        ws_itens[f'D{linha_atual}'] = item.quantidade
        ws_itens[f'E{linha_atual}'] = item.valor_unitario
        ws_itens[f'F{linha_atual}'] = item.valor_total
        ws_itens[f'G{linha_atual}'] = item.pedido.get_status_display()
        
        # Aplicar estilo
        for col in range(1, len(cabecalhos_itens) + 1):
            col_letter = get_column_letter(col)
            ws_itens[f'{col_letter}{linha_atual}'].font = normal_font
            ws_itens[f'{col_letter}{linha_atual}'].alignment = normal_alignment
            ws_itens[f'{col_letter}{linha_atual}'].border = thin_border
            
            # Formatar valores numéricos
            if col in [5, 6]:  # Colunas de valores
                ws_itens[f'{col_letter}{linha_atual}'].number_format = 'R$ #,##0.00'
        
        # Aplicar cor de fundo em linhas alternadas
        if alternar_cores and i % 2 == 1:
            for col in range(1, len(cabecalhos_itens) + 1):
                col_letter = get_column_letter(col)
                ws_itens[f'{col_letter}{linha_atual}'].fill = alternate_fill
        
        linha_atual += 1
    
    # Ajustar larguras das colunas
    for col in range(1, len(cabecalhos_itens) + 1):
        col_letter = get_column_letter(col)
        ws_itens.column_dimensions[col_letter].width = 15  # Largura padrão
    
    # Ajustes específicos
    ws_itens.column_dimensions['A'].width = 10   # ID Pedido
    ws_itens.column_dimensions['B'].width = 40   # Produto
    ws_itens.column_dimensions['G'].width = 15   # Status do Pedido
    
    # Preparar a resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=Relatorio_Pedidos_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    # Salvar o workbook para a resposta
    wb.save(response)
    
    return response

@login_required
@permission_required('pedidos.change_pedido')
def editar(request, pk):
    """Edita um pedido existente"""
    try:
        pedido = Pedido.objects.get(pk=pk)
    except Pedido.DoesNotExist:
        messages.error(request, "Pedido não encontrado.")
        return redirect('pedidos:lista')
    
    # Verificar se o pedido pode ser editado (apenas pedidos pendentes são completamente editáveis)
    pode_editar_completamente = pedido.status == 'pendente'
    
    # Verificar permissões - apenas administradores podem editar pedidos aprovados
    pode_editar_parcialmente = request.user.is_superuser or request.user.groups.filter(name='Administradores').exists()
        
    if not pode_editar_completamente and not pode_editar_parcialmente:
        messages.error(request, "Este pedido não pode ser editado pois já foi aprovado. Apenas administradores podem editar pedidos aprovados.")
        return redirect('pedidos:detalhes', pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Dados básicos do pedido que podem ser alterados
                observacoes = request.POST.get('observacoes', '')
                
                # Atualizar os dados básicos
                pedido.observacoes = observacoes
                
                # Status só pode ser alterado por administradores
                if pode_editar_parcialmente:
                    novo_status = request.POST.get('status')
                    if novo_status and novo_status != pedido.status:
                        status_anterior = pedido.status
                        pedido.status = novo_status
                        
                        # Atualiza as datas conforme o status (sem alterar a data de solicitação)
                        if novo_status == 'aprovado' and not pedido.data_aprovacao:
                            pedido.data_aprovacao = timezone.now()
                        elif novo_status == 'pedido_enviado' and not pedido.data_envio:
                            pedido.data_envio = timezone.now()
                        elif novo_status == 'entregue' and not pedido.data_entrega:
                            pedido.data_entrega = timezone.now()
                        
                        # Registrar log de alteração de status
                        PedidoLog.objects.create(
                            pedido=pedido,
                            usuario=request.user,
                            acao='alteracao_status',
                            status_anterior=status_anterior,
                            status_novo=novo_status,
                            descricao=f"Status alterado de '{dict(Pedido.STATUS_CHOICES).get(status_anterior, status_anterior)}' para '{dict(Pedido.STATUS_CHOICES).get(novo_status, novo_status)}' por {request.user.get_full_name() or request.user.username}"
                        )
                
                # Guardar alterações básicas
                pedido.save()
                
                # Processar itens do pedido apenas se o status for pendente
                if pedido.status == 'pendente':
                    # Obter IDs dos itens existentes, produtos e quantidades
                    item_ids = request.POST.getlist('item_id[]')
                    produtos_ids = request.POST.getlist('produto[]')
                    quantidades = request.POST.getlist('quantidade[]')
                    
                    # Mapear os itens existentes por ID para referência rápida
                    itens_existentes = {str(item.id): item for item in pedido.itens.all()}
                    
                    # Itens que serão mantidos (não excluídos)
                    itens_processados = set()
                    
                    # Processar itens enviados no formulário
                    for i in range(len(produtos_ids)):
                        produto_id = produtos_ids[i]
                        if not produto_id:  # Ignorar itens sem produto selecionado
                            continue
                            
                        try:
                            quantidade = int(quantidades[i])
                            if quantidade <= 0:  # Ignorar itens com quantidade inválida
                                continue
                        except (ValueError, IndexError):
                            continue
                        
                        item_id = item_ids[i] if i < len(item_ids) else ''
                        
                        # Se é um item existente
                        if item_id and item_id in itens_existentes:
                            item = itens_existentes[item_id]
                            item.produto_id = produto_id
                            item.quantidade = quantidade
                            # O valor_unitario não muda em edições
                            item.save()
                            itens_processados.add(item_id)
                        else:
                            # É um novo item
                            try:
                                produto = Produto.objects.get(id=produto_id)
                                ItemPedido.objects.create(
                                    pedido=pedido,
                                    produto=produto,
                                    quantidade=quantidade,
                                    valor_unitario=produto.valor_unitario
                                )
                            except Produto.DoesNotExist:
                                continue
                    
                    # Remover itens que não foram enviados no formulário
                    for item_id, item in itens_existentes.items():
                        if item_id not in itens_processados:
                            item.delete()
                    
                    # Verificar se o pedido ainda tem itens
                    if pedido.itens.count() == 0:
                        messages.error(request, "O pedido deve ter pelo menos um item.")
                        escolas = Escola.objects.filter(ativo=True).order_by('nome')
                        produtos = Produto.objects.filter(ativo=True).order_by('nome')
                        return render(request, 'pedidos/form.html', {
                            'pedido': pedido,
                            'escolas': escolas,
                            'produtos': produtos,
                            'status_choices': Pedido.STATUS_CHOICES
                        })
                
                # Verificar se o orçamento está dentro da validade
                if pedido.escola.budget > 0 and hasattr(pedido.escola, 'data_validade_budget') and pedido.escola.data_validade_budget:
                    data_atual = timezone.localdate()
                    if pedido.escola.data_validade_budget < data_atual:
                        messages.error(
                            request, 
                            "⚠️ ORÇAMENTO EXPIRADO! Não é possível editar pedidos neste contrato pois a data de validade do orçamento já passou."
                        )
                        escolas = Escola.objects.filter(ativo=True).order_by('nome')
                        produtos = Produto.objects.filter(ativo=True).order_by('nome')
                        return render(request, 'pedidos/form.html', {
                            'pedido': pedido,
                            'escolas': escolas,
                            'produtos': produtos,
                            'status_choices': Pedido.STATUS_CHOICES
                        })
                
                # Verificar se o valor total dos pedidos da escola não excede o orçamento
                if pedido.escola.budget > 0:
                    # Calcular o valor total dos pedidos ativos da escola (excluindo cancelados)
                    pedidos_anteriores = Pedido.objects.filter(
                        escola=pedido.escola, 
                        status__in=['pendente', 'aprovado', 'pedido_enviado', 'entregue']
                    ).exclude(id=pedido.id)  # Excluir o pedido atual
                    
                    # Calcular o valor total dos pedidos anteriores manualmente
                    valor_total_anterior = 0
                    for p in pedidos_anteriores:
                        itens_pedido = ItemPedido.objects.filter(pedido=p)
                        for item in itens_pedido:
                            valor_total_anterior += item.quantidade * item.valor_unitario
                    
                    # Calcular o valor do pedido atual
                    valor_total_atual = 0
                    itens_pedido_atual = ItemPedido.objects.filter(pedido=pedido)
                    for item in itens_pedido_atual:
                        valor_total_atual += item.quantidade * item.valor_unitario
                    
                    valor_total_com_atual = valor_total_anterior + valor_total_atual
                    
                    if valor_total_com_atual > pedido.escola.budget:
                        messages.error(
                            request, 
                            "⚠️ LIMITE DE ORÇAMENTO EXCEDIDO! Não é possível salvar as alterações pois o valor ultrapassa o orçamento disponível para este contrato."
                        )
                        escolas = Escola.objects.filter(ativo=True).order_by('nome')
                        produtos = Produto.objects.filter(ativo=True).order_by('nome')
                        return render(request, 'pedidos/form.html', {
                            'pedido': pedido,
                            'escolas': escolas,
                            'produtos': produtos,
                            'status_choices': Pedido.STATUS_CHOICES
                        })
                
                messages.success(request, "Pedido atualizado com sucesso!")
                return redirect('pedidos:detalhes', pk=pedido.pk)
                
        except Exception as e:
            messages.error(request, f"Erro ao atualizar o pedido: {str(e)}")
    
    # Carregar escolas ativas
    escolas = Escola.objects.filter(ativo=True).order_by('nome')
    
    # Carregar produtos
    produtos = Produto.objects.filter(ativo=True).order_by('nome')
    
    return render(request, 'pedidos/form.html', {
        'pedido': pedido,
        'escolas': escolas,
        'produtos': produtos,
        'status_choices': Pedido.STATUS_CHOICES
    })

@login_required
@permission_required('pedidos.change_pedido')
def atualizar_status(request, pk):
    """Atualiza o status de um pedido"""
    if request.method != 'POST':
        return redirect('pedidos:detalhes', pk=pk)
    
    pedido = get_object_or_404(Pedido, pk=pk)
    
    # Verificar se o usuário é um supervisor (Operador de Pedidos) e se tem permissão
    if request.user.groups.filter(name='Operador de Pedidos').exists():
        try:
            # Obtém o supervisor associado ao usuário logado
            supervisor = Supervisor.objects.get(usuario=request.user)
            
            # Verifica se a escola do pedido está associada ao supervisor
            if pedido.escola.supervisor != supervisor:
                messages.error(request, 'Você não tem permissão para atualizar o status deste pedido.')
                return redirect('pedidos:lista')
                
        except Supervisor.DoesNotExist:
            # Supervisor não encontrado, redireciona
            messages.warning(request, 'Sua conta de usuário não está corretamente vinculada a um supervisor. Por favor, contate o administrador.')
            return redirect('pedidos:lista')
    
    novo_status = request.POST.get('novo_status')
    
    # Validar o novo status
    status_validos = dict(Pedido.STATUS_CHOICES).keys()
    if novo_status not in status_validos:
        messages.error(request, "Status inválido.")
        return redirect('pedidos:detalhes', pk=pk)
    
    # Verificar se o usuário é um Operador de Pedidos
    is_operador = request.user.groups.filter(name='Operador de Pedidos').exists()
    
    # Verificar se o usuário é um Gerente Operacional
    is_gerente_operacional = request.user.groups.filter(name='Gerente Operacional').exists()
    
    # Verificar restrições especiais para Operador de Pedidos
    if is_operador:
        # Operador não pode aprovar pedidos
        if novo_status == 'aprovado':
            messages.error(request, "Operadores de Pedidos não têm permissão para aprovar pedidos.")
            return redirect('pedidos:detalhes', pk=pk)
        
        # Se o pedido estiver aprovado, o Operador não pode cancelá-lo
        if pedido.status != 'pendente' and novo_status == 'cancelado':
            messages.error(request, "Operadores de Pedidos só podem cancelar pedidos que ainda não foram aprovados.")
            return redirect('pedidos:detalhes', pk=pk)
    
    # Verificar se o usuário tem permissão para alterar o status
    # Apenas pedidos pendentes podem ser alterados por qualquer usuário
    # Pedidos aprovados ou em outro status só podem ser alterados por administradores ou Gerentes Operacionais
    if pedido.status != 'pendente' and not request.user.is_staff and not is_gerente_operacional:
        messages.error(request, "Apenas administradores e Gerentes Operacionais podem alterar o status de pedidos que não estão pendentes.")
        return redirect('pedidos:detalhes', pk=pk)
    
    # Verificar transições de status permitidas
    if (pedido.status == 'entregue' or pedido.status == 'cancelado') and not is_gerente_operacional and not request.user.is_superuser:
        messages.error(request, "Não é possível alterar o status de pedidos finalizados ou cancelados.")
        return redirect('pedidos:detalhes', pk=pk)
    
    # Verificar justificativa obrigatória para cancelamento
    if novo_status == 'cancelado':
        justificativa = request.POST.get('justificativa_cancelamento')
        if not justificativa or justificativa.strip() == '':
            messages.error(request, "É necessário fornecer uma justificativa para cancelar o pedido.")
            return redirect('pedidos:detalhes', pk=pk)
        pedido.justificativa_cancelamento = justificativa
    
    # Verificar informações obrigatórias para entrega
    data_entrega_manual = None
    recebido_por = None
    
    if novo_status == 'entregue':
        data_entrega = request.POST.get('data_entrega')
        hora_entrega = request.POST.get('hora_entrega')
        recebido_por = request.POST.get('recebido_por')
        
        # Validar os campos obrigatórios
        if not data_entrega or not hora_entrega or not recebido_por:
            messages.error(request, "Para marcar como entregue, é obrigatório informar a data, hora e quem recebeu o pedido.")
            return redirect('pedidos:detalhes', pk=pk)
        
        # Converter data e hora para um objeto datetime
        try:
            data_parts = data_entrega.split('-')
            hora_parts = hora_entrega.split(':')
            
            ano = int(data_parts[0])
            mes = int(data_parts[1])
            dia = int(data_parts[2])
            hora = int(hora_parts[0])
            minuto = int(hora_parts[1])
            
            data_entrega_manual = timezone.make_aware(datetime(ano, mes, dia, hora, minuto))
            
            # Verificar se a data não está no futuro
            if data_entrega_manual > timezone.now():
                messages.error(request, "A data e hora da entrega não podem estar no futuro.")
                return redirect('pedidos:detalhes', pk=pk)
                
        except (ValueError, IndexError):
            messages.error(request, "Formato de data ou hora inválido.")
            return redirect('pedidos:detalhes', pk=pk)
    
    # Verificar progressão lógica de status
    status_ordem = {
        'pendente': 0,
        'aprovado': 1,
        'pedido_enviado': 2,
        'entregue': 3
    }
    
    # Impedir retrocesso de status (exceto para cancelamento)
    if novo_status != 'cancelado' and status_ordem.get(novo_status, 0) < status_ordem.get(pedido.status, 0):
        messages.error(request, "Não é possível retroceder o status do pedido.")
        return redirect('pedidos:detalhes', pk=pk)
    
    # Guardar status anterior para o log
    status_anterior = pedido.status
    status_anterior_display = dict(Pedido.STATUS_CHOICES)[status_anterior]
    status_novo_display = dict(Pedido.STATUS_CHOICES)[novo_status]
    
    # Atualizar o status
    try:
        pedido.atualizar_status(novo_status, data_entrega_manual, recebido_por)
        
        # Registrar log de alteração de status
        descricao = f"Status alterado de '{status_anterior_display}' para '{status_novo_display}' por {request.user.get_full_name() or request.user.username}"
        
        # Adicionar informações sobre a entrega ao log, se aplicável
        if novo_status == 'entregue' and recebido_por:
            descricao += f". Recebido por: {recebido_por} em {data_entrega_manual.strftime('%d/%m/%Y %H:%M')}"
        
        PedidoLog.objects.create(
            pedido=pedido,
            usuario=request.user,
            acao='alteracao_status',
            status_anterior=status_anterior,
            status_novo=novo_status,
            descricao=descricao
        )
        
        # Mensagem personalizada com base no status
        if novo_status == 'aprovado':
            messages.success(request, "Pedido aprovado com sucesso! A partir de agora, apenas administradores poderão alterá-lo.")
        elif novo_status == 'cancelado':
            messages.success(request, "Pedido cancelado com sucesso.")
        else:
            status_nome = dict(Pedido.STATUS_CHOICES)[novo_status]
            messages.success(request, f"Status do pedido atualizado para '{status_nome}' com sucesso.")
            
    except Exception as e:
        messages.error(request, f"Erro ao atualizar status: {str(e)}")
    
    return redirect('pedidos:detalhes', pk=pk)
