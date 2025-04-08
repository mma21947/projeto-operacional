from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from .models import Escola, Supervisor
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.db import transaction
import tempfile
import os
from django.contrib.auth.decorators import login_required, permission_required
from datetime import datetime
import qrcode
import base64
from io import BytesIO

# Lista de estados brasileiros
ESTADOS_BRASILEIROS = [
    ('AC', 'Acre'),
    ('AL', 'Alagoas'),
    ('AP', 'Amapá'),
    ('AM', 'Amazonas'),
    ('BA', 'Bahia'),
    ('CE', 'Ceará'),
    ('DF', 'Distrito Federal'),
    ('ES', 'Espírito Santo'),
    ('GO', 'Goiás'),
    ('MA', 'Maranhão'),
    ('MT', 'Mato Grosso'),
    ('MS', 'Mato Grosso do Sul'),
    ('MG', 'Minas Gerais'),
    ('PA', 'Pará'),
    ('PB', 'Paraíba'),
    ('PR', 'Paraná'),
    ('PE', 'Pernambuco'),
    ('PI', 'Piauí'),
    ('RJ', 'Rio de Janeiro'),
    ('RN', 'Rio Grande do Norte'),
    ('RS', 'Rio Grande do Sul'),
    ('RO', 'Rondônia'),
    ('RR', 'Roraima'),
    ('SC', 'Santa Catarina'),
    ('SP', 'São Paulo'),
    ('SE', 'Sergipe'),
    ('TO', 'Tocantins'),
]

@login_required
@permission_required('escolas.view_escola')
def lista(request):
    """Exibe a lista de escolas"""
    
    # Verifica se o usuário é um supervisor (grupo 'Operador de Pedidos')
    if request.user.groups.filter(name='Operador de Pedidos').exists():
        try:
            # Tenta obter o supervisor associado ao usuário logado
            supervisor = Supervisor.objects.get(usuario=request.user)
            # Filtra apenas as escolas associadas a este supervisor
            escolas = Escola.objects.filter(supervisor=supervisor)
            
            # Mensagem de contexto para o usuário
            messages.info(request, f'Você está visualizando apenas os contratos sob sua supervisão.')
            
        except Supervisor.DoesNotExist:
            # Supervisor não encontrado, não mostra escolas
            escolas = Escola.objects.none()
            messages.warning(request, 'Sua conta de usuário não está corretamente vinculada a um supervisor. Por favor, contate o administrador.')
    else:
        # Usuários administradores ou outros veem todas as escolas
        escolas = Escola.objects.all()
    
    return render(request, 'escolas/lista.html', {'escolas': escolas})

def novo(request):
    """Adiciona uma nova escola"""
    if request.method == 'POST':
        # Captura dados do formulário
        nome = request.POST.get('nome')
        codigo = request.POST.get('codigo')
        lote = request.POST.get('lote')
        empresa = request.POST.get('empresa')
        budget = request.POST.get('budget', 0)
        endereco = request.POST.get('endereco')
        cep = request.POST.get('cep')
        cidade = request.POST.get('cidade')
        estado = request.POST.get('estado')
        telefone = request.POST.get('telefone')
        email = request.POST.get('email')
        supervisor_id = request.POST.get('supervisor')
        ativo = request.POST.get('ativo') == 'on'
        
        # Validações
        if not nome:
            messages.error(request, 'O nome da escola é obrigatório.')
            supervisores = Supervisor.objects.filter(ativo=True).order_by('nome')
            return render(request, 'escolas/form.html', {
                'request': request,
                'supervisores': supervisores,
                'estados': ESTADOS_BRASILEIROS
            })
        
        # Conversão do orçamento
        try:
            budget = float(budget) if budget else 0
        except ValueError:
            budget = 0
        
        # Data de validade do orçamento
        data_validade_budget = request.POST.get('data_validade_budget')
        if data_validade_budget:
            try:
                # Convertendo de string para data
                data_validade = datetime.strptime(data_validade_budget, '%Y-%m-%d').date()
            except ValueError:
                data_validade = None
        else:
            data_validade = None
        
        # Cria nova escola
        escola = Escola(
            nome=nome,
            codigo=codigo,
            lote=lote,
            empresa=empresa,
            budget=budget,
            data_validade_budget=data_validade,
            endereco=endereco,
            cep=cep,
            cidade=cidade,
            estado=estado,
            telefone=telefone,
            email=email,
            ativo=ativo
        )
        
        # Associa o supervisor se fornecido
        if supervisor_id:
            try:
                supervisor = Supervisor.objects.get(pk=supervisor_id)
                escola.supervisor = supervisor
            except Supervisor.DoesNotExist:
                pass
        
        escola.save()
        messages.success(request, f'Escola "{nome}" criada com sucesso!')
        return redirect('escolas:lista')
    
    # Para requisições GET, carrega os supervisores para o select
    supervisores = Supervisor.objects.filter(ativo=True).order_by('nome')
    return render(request, 'escolas/form.html', {
        'supervisores': supervisores,
        'estados': ESTADOS_BRASILEIROS
    })

def editar(request, pk):
    """Edita uma escola existente"""
    escola = get_object_or_404(Escola, pk=pk)
    
    if request.method == 'POST':
        # Captura dados do formulário
        escola.nome = request.POST.get('nome')
        escola.codigo = request.POST.get('codigo')
        escola.lote = request.POST.get('lote')
        escola.empresa = request.POST.get('empresa')
        
        # Conversão do orçamento
        budget = request.POST.get('budget', 0)
        try:
            escola.budget = float(budget) if budget else 0
        except ValueError:
            escola.budget = 0
            
        # Data de validade do orçamento
        data_validade_budget = request.POST.get('data_validade_budget')
        if data_validade_budget:
            try:
                # Convertendo de string para data
                escola.data_validade_budget = datetime.strptime(data_validade_budget, '%Y-%m-%d').date()
            except ValueError:
                escola.data_validade_budget = None
        else:
            escola.data_validade_budget = None
            
        escola.endereco = request.POST.get('endereco')
        escola.cep = request.POST.get('cep')
        escola.cidade = request.POST.get('cidade')
        escola.estado = request.POST.get('estado')
        escola.telefone = request.POST.get('telefone')
        escola.email = request.POST.get('email')
        escola.ativo = request.POST.get('ativo') == 'on'
        
        # Associa o supervisor se fornecido
        supervisor_id = request.POST.get('supervisor')
        if supervisor_id:
            try:
                supervisor = Supervisor.objects.get(pk=supervisor_id)
                escola.supervisor = supervisor
            except Supervisor.DoesNotExist:
                escola.supervisor = None
        else:
            escola.supervisor = None
        
        # Validações
        if not escola.nome:
            messages.error(request, 'O nome da escola é obrigatório.')
            supervisores = Supervisor.objects.filter(ativo=True).order_by('nome')
            return render(request, 'escolas/form.html', {
                'escola': escola, 
                'supervisores': supervisores,
                'estados': ESTADOS_BRASILEIROS
            })
            
        escola.save()
        messages.success(request, f'Escola "{escola.nome}" atualizada com sucesso!')
        return redirect('escolas:detalhes', pk=escola.pk)
    
    # Para requisições GET
    supervisores = Supervisor.objects.filter(ativo=True).order_by('nome')
    return render(request, 'escolas/form.html', {
        'escola': escola, 
        'supervisores': supervisores,
        'estados': ESTADOS_BRASILEIROS
    })

@login_required
@permission_required('escolas.view_escola')
def detalhes(request, pk):
    """Exibe detalhes de uma escola"""
    escola = get_object_or_404(Escola, pk=pk)
    
    # Verifica se o usuário é um supervisor (grupo 'Operador de Pedidos')
    if request.user.groups.filter(name='Operador de Pedidos').exists():
        try:
            # Tenta obter o supervisor associado ao usuário logado
            supervisor = Supervisor.objects.get(usuario=request.user)
            
            # Verifica se a escola pertence ao supervisor
            if escola.supervisor != supervisor:
                messages.error(request, 'Você não tem permissão para acessar os detalhes deste contrato.')
                return redirect('escolas:lista')
                
        except Supervisor.DoesNotExist:
            # Supervisor não encontrado, redireciona
            messages.warning(request, 'Sua conta de usuário não está corretamente vinculada a um supervisor. Por favor, contate o administrador.')
            return redirect('escolas:lista')
    
    # Gerar QR code para WhatsApp se houver supervisor
    supervisor_qr_code = None
    whatsapp_link = None
    
    if escola.supervisor and escola.supervisor.telefone:
        # Formatar o número de telefone para padrão internacional
        # Remover caracteres não numéricos
        telefone_limpo = ''.join(filter(str.isdigit, escola.supervisor.telefone))
        
        # Se o número já começar com +55 (Brasil), usar como está
        # Caso contrário, adicionar o código do país (+55 para Brasil)
        if not telefone_limpo.startswith('55'):
            telefone_limpo = '55' + telefone_limpo
            
        # Criar link do WhatsApp (formato: https://wa.me/5511999999999)
        whatsapp_link = f"https://wa.me/{telefone_limpo}"
        
        # Gerar QR code para o link do WhatsApp
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(whatsapp_link)
        qr.make(fit=True)
        
        # Criar imagem e converter para base64
        img = qr.make_image(fill_color="black", back_color="white")
        buffer = BytesIO()
        img.save(buffer, format="PNG")
        supervisor_qr_code = base64.b64encode(buffer.getvalue()).decode("utf-8")
    
    # Buscar pedidos associados à escola (se o app 'pedidos' estiver instalado)
    pedidos = []
    try:
        from pedidos.models import Pedido
        pedidos = Pedido.objects.filter(escola=escola).order_by('-data_solicitacao')
    except ImportError:
        # App 'pedidos' não está instalado
        pass
    
    return render(request, 'escolas/detalhes.html', {
        'escola': escola,
        'pedidos': pedidos,
        'supervisor_qr_code': supervisor_qr_code,
        'whatsapp_link': whatsapp_link,
        'tem_endereco': bool(escola.endereco or escola.cidade)
    })

def importar(request):
    """Importa escolas de um arquivo Excel"""
    if request.method == 'POST' and request.FILES.get('arquivo'):
        arquivo = request.FILES['arquivo']
        importar_supervisores = request.POST.get('importar_supervisores') == 'on'
        
        # Verifica extensão do arquivo
        if not arquivo.name.endswith(('.xls', '.xlsx')):
            messages.error(request, 'Formato de arquivo inválido. Apenas arquivos .xls e .xlsx são permitidos.')
            return render(request, 'escolas/importar.html')
        
        try:
            # Lê a planilha Excel usando o pandas
            df = pd.read_excel(arquivo)
            
            # Verifica se o DataFrame está vazio
            if df.empty:
                messages.error(request, 'O arquivo está vazio ou não contém dados válidos.')
                return render(request, 'escolas/importar.html')
            
            # Verifica se as colunas necessárias existem
            colunas_obrigatorias = ["Nome da Escola"]
            for coluna in colunas_obrigatorias:
                if coluna not in df.columns:
                    messages.error(request, f'Coluna obrigatória "{coluna}" não encontrada na planilha.')
                    return render(request, 'escolas/importar.html')
            
            # Mapeia as colunas do Excel para os campos do modelo
            mapeamento = {
                "Nome da Escola": "nome",
                "Código": "codigo",
                "Lote": "lote",
                "Empresa": "empresa",
                "Endereço": "endereco",
                "CEP": "cep",
                "Cidade": "cidade",
                "Estado": "estado",
                "Telefone": "telefone",
                "Email": "email"
            }
            
            # Contadores para feedback
            total_criadas = 0
            total_atualizadas = 0
            
            # Processa cada linha da planilha
            with transaction.atomic():
                for _, row in df.iterrows():
                    # Verifica se o nome da escola está preenchido
                    if pd.isna(row["Nome da Escola"]) or not str(row["Nome da Escola"]).strip():
                        continue
                    
                    # Dados da escola
                    escola_data = {}
                    for excel_col, model_field in mapeamento.items():
                        if excel_col in df.columns and not pd.isna(row[excel_col]):
                            escola_data[model_field] = str(row[excel_col]).strip()
                    
                    # Verifica se já existe uma escola com o mesmo código
                    codigo = escola_data.get('codigo')
                    escola = None
                    
                    if codigo:
                        try:
                            escola = Escola.objects.get(codigo=codigo)
                            # Atualiza a escola existente
                            for field, value in escola_data.items():
                                setattr(escola, field, value)
                            escola.save()
                            total_atualizadas += 1
                        except Escola.DoesNotExist:
                            # Escola não existe, será criada mais abaixo
                            pass
                    
                    # Se a escola não foi encontrada pelo código, cria uma nova
                    if not escola:
                        escola = Escola.objects.create(**escola_data)
                        total_criadas += 1
                    
                    # Processa o supervisor, se a opção estiver habilitada
                    if importar_supervisores and "Nome do Supervisor" in df.columns and not pd.isna(row["Nome do Supervisor"]):
                        supervisor_nome = str(row["Nome do Supervisor"]).strip()
                        supervisor_email = str(row["Email do Supervisor"]).strip() if "Email do Supervisor" in df.columns and not pd.isna(row["Email do Supervisor"]) else None
                        supervisor_telefone = str(row["Telefone do Supervisor"]).strip() if "Telefone do Supervisor" in df.columns and not pd.isna(row["Telefone do Supervisor"]) else None
                        
                        # Tenta encontrar o supervisor por email (se fornecido)
                        supervisor = None
                        if supervisor_email:
                            supervisor = Supervisor.objects.filter(email=supervisor_email).first()
                        
                        # Se não encontrou por email, tenta por nome
                        if not supervisor:
                            supervisor = Supervisor.objects.filter(nome=supervisor_nome).first()
                        
                        # Se não encontrou, cria um novo supervisor
                        if not supervisor:
                            supervisor = Supervisor.objects.create(
                                nome=supervisor_nome,
                                email=supervisor_email,
                                telefone=supervisor_telefone
                            )
                        
                        # Associa o supervisor à escola
                        escola.supervisor = supervisor
                        escola.save()
            
            # Mensagem de sucesso com os detalhes
            if total_criadas > 0 or total_atualizadas > 0:
                mensagem = f'Importação concluída com sucesso! '
                if total_criadas > 0:
                    mensagem += f'{total_criadas} escolas criadas. '
                if total_atualizadas > 0:
                    mensagem += f'{total_atualizadas} escolas atualizadas.'
                messages.success(request, mensagem)
            else:
                messages.warning(request, 'Nenhuma escola foi importada. Verifique o arquivo e tente novamente.')
            
            return redirect('escolas:lista')
            
        except Exception as e:
            messages.error(request, f'Erro ao processar o arquivo: {str(e)}')
            return render(request, 'escolas/importar.html')
    
    return render(request, 'escolas/importar.html')

def download_modelo(request):
    """Download do modelo Excel para importação de escolas"""
    # Cria uma nova planilha
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Modelo Importação Escolas"
    
    # Define os cabeçalhos
    cabecalhos = [
        "Nome da Escola", "Código", "Lote", "Empresa", "Endereço", "CEP", "Cidade", "Estado",
        "Telefone", "Email", "Nome do Supervisor", "Email do Supervisor", "Telefone do Supervisor"
    ]
    
    for col_num, cabecalho in enumerate(cabecalhos, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = cabecalho
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Exemplo de dados
    exemplo = [
        "Escola Municipal José Silva", "ESC001", "Lote 3", "Empresa ABC", "Rua das Flores, 123", "12345-678",
        "São Paulo", "SP", "(11) 1234-5678", "contato@escolajosesilva.edu.br",
        "Carlos Souza", "carlos@exemplo.com", "(11) 98765-4321"
    ]
    
    for col_num, valor in enumerate(exemplo, 1):
        sheet.cell(row=2, column=col_num).value = valor
    
    # Ajusta a largura das colunas
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width
    
    # Cria a resposta HTTP com o arquivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=modelo_escolas.xlsx'
    
    # Salva o workbook na resposta
    workbook.save(response)
    
    return response
