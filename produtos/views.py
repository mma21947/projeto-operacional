from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from .models import Produto
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.db import transaction
from decimal import Decimal

def lista(request):
    """Exibe a lista de produtos"""
    produtos = Produto.objects.all()
    return render(request, 'produtos/lista.html', {'produtos': produtos})

def novo(request):
    """Adiciona um novo produto"""
    if request.method == 'POST':
        # Extrai os dados do formulário
        nome = request.POST.get('nome')
        descricao = request.POST.get('descricao')
        valor_unitario = request.POST.get('valor_unitario').replace(',', '.')
        unidade_medida = request.POST.get('unidade_medida')
        codigo = request.POST.get('codigo')
        
        # Valida os dados
        if not nome:
            messages.error(request, 'O nome do produto é obrigatório.')
            return render(request, 'produtos/form.html', {'produto': request.POST})
        
        try:
            valor_unitario = Decimal(valor_unitario)
            if valor_unitario <= 0:
                raise ValueError("Valor deve ser maior que zero")
        except:
            messages.error(request, 'O valor unitário deve ser um número positivo.')
            return render(request, 'produtos/form.html', {'produto': request.POST})
        
        # Cria o produto
        produto = Produto(
            nome=nome,
            descricao=descricao,
            valor_unitario=valor_unitario,
            unidade_medida=unidade_medida,
            codigo=codigo,
        )
        produto.save()
        
        messages.success(request, 'Produto adicionado com sucesso!')
        return redirect('produtos:lista')
    
    return render(request, 'produtos/form.html')

def editar(request, pk):
    """Edita um produto existente"""
    produto = get_object_or_404(Produto, pk=pk)
    
    if request.method == 'POST':
        # Extrai os dados do formulário
        produto.nome = request.POST.get('nome')
        produto.descricao = request.POST.get('descricao')
        produto.valor_unitario = request.POST.get('valor_unitario').replace(',', '.')
        produto.unidade_medida = request.POST.get('unidade_medida')
        produto.codigo = request.POST.get('codigo')
        produto.ativo = 'ativo' in request.POST
        
        # Valida os dados
        if not produto.nome:
            messages.error(request, 'O nome do produto é obrigatório.')
            return render(request, 'produtos/form.html', {'produto': produto})
        
        try:
            produto.valor_unitario = Decimal(produto.valor_unitario)
            if produto.valor_unitario <= 0:
                raise ValueError("Valor deve ser maior que zero")
        except:
            messages.error(request, 'O valor unitário deve ser um número positivo.')
            return render(request, 'produtos/form.html', {'produto': produto})
        
        # Salva o produto
        produto.save()
        
        messages.success(request, 'Produto atualizado com sucesso!')
        return redirect('produtos:lista')
    
    return render(request, 'produtos/form.html', {'produto': produto})

def importar(request):
    """Importa produtos de um arquivo Excel"""
    if request.method == 'POST' and 'arquivo' in request.FILES:
        arquivo = request.FILES['arquivo']
        
        # Verifica a extensão do arquivo
        if not arquivo.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Formato de arquivo não suportado. Por favor, envie um arquivo Excel (.xlsx ou .xls).')
            return redirect('produtos:importar')
        
        try:
            # Carrega o arquivo Excel
            df = pd.read_excel(arquivo)
            
            # Verifica se as colunas necessárias existem
            colunas_necessarias = ['Nome do Produto', 'Valor Unitário']
            for coluna in colunas_necessarias:
                if coluna not in df.columns:
                    messages.error(request, f'Coluna "{coluna}" não encontrada no arquivo. Verifique o formato do arquivo.')
                    return redirect('produtos:importar')
            
            # Mapeia os nomes das colunas no Excel para os campos do modelo
            mapeamento = {
                'Nome do Produto': 'nome',
                'Descrição': 'descricao',
                'Valor Unitário': 'valor_unitario',
                'Unidade de Medida': 'unidade_medida',
                'Código': 'codigo',
            }
            
            produtos_importados = 0
            produtos_atualizados = 0
            
            # Inicia uma transação para garantir consistência
            with transaction.atomic():
                for _, row in df.iterrows():
                    dados_produto = {}
                    
                    # Extrai os dados da linha com base no mapeamento
                    for coluna_excel, campo_modelo in mapeamento.items():
                        if coluna_excel in df.columns and not pd.isna(row[coluna_excel]):
                            dados_produto[campo_modelo] = row[coluna_excel]
                    
                    # Verifica campos obrigatórios
                    if 'nome' not in dados_produto or 'valor_unitario' not in dados_produto:
                        continue
                    
                    # Verifica se o produto já existe (por código)
                    produto_existente = None
                    if 'codigo' in dados_produto and dados_produto['codigo']:
                        try:
                            produto_existente = Produto.objects.get(codigo=dados_produto['codigo'])
                        except Produto.DoesNotExist:
                            pass
                    
                    if produto_existente:
                        # Atualiza o produto existente
                        for campo, valor in dados_produto.items():
                            setattr(produto_existente, campo, valor)
                        produto_existente.save()
                        produtos_atualizados += 1
                    else:
                        # Cria um novo produto
                        # Define a unidade de medida padrão se não informada
                        if 'unidade_medida' not in dados_produto:
                            dados_produto['unidade_medida'] = 'Unidade'
                        
                        Produto.objects.create(**dados_produto)
                        produtos_importados += 1
            
            messages.success(request, f'Importação concluída: {produtos_importados} produtos importados e {produtos_atualizados} produtos atualizados.')
            return redirect('produtos:lista')
            
        except Exception as e:
            messages.error(request, f'Erro ao processar o arquivo: {str(e)}')
            return redirect('produtos:importar')
    
    return render(request, 'produtos/importar.html')

def download_modelo(request):
    """Download do modelo Excel para importação de produtos"""
    # Cria uma nova planilha
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Modelo Importação Produtos"
    
    # Define os cabeçalhos
    cabecalhos = ["Nome do Produto", "Descrição", "Valor Unitário", "Unidade de Medida", "Código"]
    for col_num, cabecalho in enumerate(cabecalhos, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = cabecalho
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Adiciona exemplos
    exemplos = [
        ["Caderno Espiral", "Caderno espiral 200 folhas, capa dura", 15.90, "Unidade", "CAD001"],
        ["Lápis HB", "Caixa com 12 lápis pretos HB", 12.50, "Caixa", "LAP002"],
    ]
    
    for row_num, exemplo in enumerate(exemplos, 2):
        for col_num, valor in enumerate(exemplo, 1):
            sheet.cell(row=row_num, column=col_num).value = valor
    
    # Ajusta a largura das colunas
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width
    
    # Cria a resposta HTTP com o arquivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=modelo_produtos.xlsx'
    
    # Salva o workbook na resposta
    workbook.save(response)
    
    return response
